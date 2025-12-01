from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie
from django.http import JsonResponse, HttpResponse
from django.utils.timezone import localtime, now, make_aware
from django.conf import settings

from .models import ReceivedMaterial
from .utils import business_day, get_current_business_day

import json
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta, timezone as dt_timezone
from io import BytesIO
import os

import openpyxl
from openpyxl.styles import Alignment, Font


# =====================
# Helper Functions
# =====================

def _parse_iso_local(s: str):
    """
    Parses ISO string 'YYYY-MM-DDTHH:MM:SS' (or variants) as local time.
    Returns timezone-aware datetime in project's timezone.
    """
    s = (s or "").strip().replace(" ", "T")

    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s, fmt)
            break
        except ValueError:
            continue
    else:
        raise ValueError(f"Unsupported datetime format: {s}")

    # Make timezone-aware in project's timezone
    if dt.tzinfo is None:
        return make_aware(dt)
    return dt


def _get_date_range_for_period(period: str):
    """
    Returns (start_dt, end_dt) for a given period ('today', 'prev').
    Both are timezone-aware datetimes.
    """
    local_now = localtime(now())
    today_date = local_now.date()

    if period == 'today':
        # From midnight today to now
        start = make_aware(datetime.combine(today_date, datetime.min.time()))
        end = local_now
    elif period == 'prev':
        # Yesterday: from midnight to 23:59:59
        yesterday = today_date - timedelta(days=1)
        start = make_aware(datetime.combine(yesterday, datetime.min.time()))
        end = make_aware(datetime.combine(yesterday, datetime.max.time()))
    else:
        # Default to today
        start = make_aware(datetime.combine(today_date, datetime.min.time()))
        end = local_now

    return start, end


def _safe_decimal(value, default='0'):
    """Safely converts value to Decimal with error handling."""
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)


# =====================
# Views
# =====================

@ensure_csrf_cookie
@login_required
def home(request):
    """
    Main scales home page.
    Shows today's records and provides manager view for history.
    """
    is_manager = request.session.get("user_role") == "managers"
    company = request.session.get("company_slug")

    # Get today's and yesterday's records
    start_today, end_today = _get_date_range_for_period('today')
    start_prev, end_prev = _get_date_range_for_period('prev')

    qs = ReceivedMaterial.objects.all()
    if company:
        qs = qs.filter(company_slug=company)

    received_today = qs.filter(
        created_at__gte=start_today,
        created_at__lt=end_today
    ).order_by("-created_at")

    received_prev = qs.filter(
        created_at__gte=start_prev,
        created_at__lt=end_prev
    ).order_by("-created_at")

    return render(request, "scales/home.html", {
        "title": "Scales • Home",
        "is_manager": is_manager,
        "start": start_today,
        "end": end_today,
        "pstart": start_prev,
        "pend": end_prev,
        "received_today": received_today,
        "received_prev": received_prev,
    })


@login_required
@require_http_methods(["GET"])
def api_list_received(request):
    """
    GET /scales/api/received/

    Query params:
        - period=today|prev|all
        - from=YYYY-MM-DDTHH:MM:SS&to=YYYY-MM-DDTHH:MM:SS (local time)

    Returns: {"items": [...]}
    Each item includes:
        - id, date (ISO), report_day (YYYY-MM-DD), material, gross, net, supplier, tag
    """
    company = request.session.get("company_slug")
    period = (request.GET.get("period") or "").lower()
    frm = request.GET.get("from")
    to = request.GET.get("to")

    qs = ReceivedMaterial.objects.all()
    if company:
        qs = qs.filter(company_slug=company)

    # Filter by period
    if period in ("today", "prev"):
        start, end = _get_date_range_for_period(period)
        qs = qs.filter(created_at__gte=start, created_at__lt=end)

    # Filter by custom date range
    elif frm and to:
        try:
            from_dt = _parse_iso_local(frm)
            to_dt = _parse_iso_local(to)
            qs = qs.filter(created_at__gte=from_dt, created_at__lte=to_dt)
        except ValueError as e:
            return JsonResponse({
                "items": [],
                "error": str(e)
            }, status=400)

    # period=all or default to today
    elif period != "all":
        start, end = _get_date_range_for_period('today')
        qs = qs.filter(created_at__gte=start, created_at__lt=end)

    # Limit results
    limit = 10000 if (frm and to) or period == "all" else 500
    qs = qs.order_by("-created_at")[:limit]

    # Build response
    items = []
    for r in qs:
        created_local = localtime(r.created_at)
        rep_day = r.report_day or business_day(created_local)

        items.append({
            "id": r.id,
            "date": created_local.isoformat(timespec="seconds"),
            "report_day": rep_day.isoformat(),
            "material": r.material,
            "gross": float(r.gross_kg),
            "net": float(r.net_kg),
            "supplier": r.supplier,
            "tag": r.tag,
        })

    return JsonResponse({"items": items})


@login_required
@csrf_protect
@require_http_methods(["POST"])
def api_create_received(request):
    """
    POST /scales/api/received/create/
    Body: {"material", "gross_kg", "net_kg", "supplier", "tag"}

    Creates new ReceivedMaterial record.
    """
    try:
        body = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    # Validate required fields
    required = ["material", "supplier", "tag"]
    missing = [f for f in required if not body.get(f)]
    if missing:
        return JsonResponse({
            "error": f"Missing required fields: {', '.join(missing)}"
        }, status=400)

    # Get company
    company = request.session.get("company_slug") or "unknown"

    # Parse weights with error handling
    try:
        gross_kg = _safe_decimal(body.get("gross_kg") or body.get("gross"))
        net_kg = _safe_decimal(body.get("net_kg") or body.get("net"))

        if gross_kg <= 0 or net_kg <= 0:
            raise ValueError("Weights must be positive")
    except (ValueError, InvalidOperation) as e:
        return JsonResponse({
            "error": f"Invalid weight values: {e}"
        }, status=400)

    # Create record
    item = ReceivedMaterial.objects.create(
        material=body["material"].strip(),
        gross_kg=gross_kg,
        net_kg=net_kg,
        supplier=body["supplier"].strip(),
        tag=str(body["tag"]).strip(),
        company_slug=company,
        created_by=request.user if request.user.is_authenticated else None,
    )

    # Return created item
    created_local = localtime(item.created_at)
    rep_day = item.report_day or business_day(created_local)

    return JsonResponse({
        "ok": True,
        "item": {
            "id": item.id,
            "date": created_local.isoformat(timespec="seconds"),
            "report_day": rep_day.isoformat(),
            "material": item.material,
            "gross": float(item.gross_kg),
            "net": float(item.net_kg),
            "supplier": item.supplier,
            "tag": item.tag,
        }
    })


@login_required
@csrf_protect
@require_http_methods(["POST"])
def api_update_received(request, pk):
    """
    POST /scales/api/received/<pk>/update/
    Body: {"material", "gross_kg", "net_kg", "supplier", "tag"}

    Updates existing ReceivedMaterial record.
    """
    company = request.session.get("company_slug")
    obj = get_object_or_404(ReceivedMaterial, pk=pk, company_slug=company)

    try:
        body = json.loads(request.body.decode("utf-8"))
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    # Update fields with validation
    if "material" in body:
        obj.material = body["material"].strip()

    if "gross_kg" in body or "gross" in body:
        try:
            obj.gross_kg = _safe_decimal(body.get("gross_kg") or body.get("gross"))
        except (ValueError, InvalidOperation):
            return JsonResponse({"error": "Invalid gross weight"}, status=400)

    if "net_kg" in body or "net" in body:
        try:
            obj.net_kg = _safe_decimal(body.get("net_kg") or body.get("net"))
        except (ValueError, InvalidOperation):
            return JsonResponse({"error": "Invalid net weight"}, status=400)

    if "supplier" in body:
        obj.supplier = body["supplier"].strip()

    if "tag" in body:
        obj.tag = str(body["tag"]).strip()

    obj.save()

    # Return updated item
    created_local = localtime(obj.created_at)
    rep_day = obj.report_day or business_day(created_local)

    return JsonResponse({
        "ok": True,
        "item": {
            "id": obj.id,
            "date": created_local.isoformat(timespec="seconds"),
            "report_day": rep_day.isoformat(),
            "material": obj.material,
            "gross": float(obj.gross_kg),
            "net": float(obj.net_kg),
            "supplier": obj.supplier,
            "tag": obj.tag,
        }
    })


@login_required
@csrf_protect
@require_http_methods(["POST"])
def api_delete_received(request, pk):
    """
    POST /scales/api/received/<pk>/delete/

    Deletes ReceivedMaterial record.
    """
    company = request.session.get("company_slug")
    obj = get_object_or_404(ReceivedMaterial, pk=pk, company_slug=company)
    obj.delete()

    return JsonResponse({"ok": True})


# =====================
# Export Views
# =====================

@login_required
def export_daily_pdf(request):
    """
    Daily Material Report (Scales) — PDF with pagination.
    Query: ?date=YYYY-MM-DD (defaults to today)
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader

    # Get parameters
    company_slug = request.session.get("company_slug") or "unknown"
    company_name = request.session.get("company_name") or company_slug
    date_str = request.GET.get("date") or datetime.now().strftime("%Y-%m-%d")

    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Get records for the date
    qs = (
        ReceivedMaterial.objects
        .filter(company_slug=company_slug, report_day=date_obj)
        .order_by("created_at")
    )

    # Build main table data
    data = [["Material", "Gross (kg)", "Net (kg)", "Supplier", "Tag #"]]
    total_gross = 0.0
    total_net = 0.0

    # Summaries
    by_material = {}
    by_supplier_mat = {}

    for r in qs:
        g = float(r.gross_kg)
        n = float(r.net_kg)

        data.append([
            r.material,
            f"{g:.1f}",
            f"{n:.1f}",
            r.supplier,
            str(r.tag),
        ])

        total_gross += g
        total_net += n

        # By material
        if r.material not in by_material:
            by_material[r.material] = {"gross": 0.0, "net": 0.0, "count": 0}
        by_material[r.material]["gross"] += g
        by_material[r.material]["net"] += n
        by_material[r.material]["count"] += 1

        # By supplier and material
        if r.supplier not in by_supplier_mat:
            by_supplier_mat[r.supplier] = {"_subtotal": {"gross": 0.0, "net": 0.0}}

        sup_map = by_supplier_mat[r.supplier]
        if r.material not in sup_map:
            sup_map[r.material] = {"gross": 0.0, "net": 0.0}

        sup_map[r.material]["gross"] += g
        sup_map[r.material]["net"] += n
        sup_map["_subtotal"]["gross"] += g
        sup_map["_subtotal"]["net"] += n

    # PDF setup
    buffer = BytesIO()
    pdf_canvas = canvas.Canvas(buffer, pagesize=A4)
    PAGE_W, PAGE_H = A4
    M_L, M_R, M_T, M_B = 30, 30, 40, 40
    usable_w = PAGE_W - M_L - M_R

    logo_path = os.path.join(settings.BASE_DIR, 'crm', 'static', 'crm', 'images', 'logo3.png')

    def draw_header():
        y = PAGE_H - 60
        if os.path.exists(logo_path):
            pdf_canvas.drawImage(ImageReader(logo_path), M_L, y - 50, width=70, height=70, mask='auto')

        pdf_canvas.setFont("Helvetica-Bold", 16)
        pdf_canvas.drawCentredString(PAGE_W / 2, y, "Daily Material Report")

        pdf_canvas.setFont("Helvetica-Bold", 11)
        pdf_canvas.setFillColor(colors.darkblue)
        pdf_canvas.drawRightString(PAGE_W - M_R, y, company_name)

        pdf_canvas.setFont("Helvetica", 9)
        pdf_canvas.setFillColor(colors.HexColor("#555555"))
        pdf_canvas.drawRightString(PAGE_W - M_R, y - 14, f"Date: {date_obj.strftime('%d %b %Y')}")

        pdf_canvas.setStrokeColor(colors.HexColor("#aaaaaa"))
        pdf_canvas.setLineWidth(0.5)
        pdf_canvas.line(M_L, y - 50, PAGE_W - M_R, y - 50)

        return PAGE_H - 135

    def table_style():
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ])

    def ensure_space(y_cur, need=40):
        if y_cur - need < M_B:
            pdf_canvas.showPage()
            return draw_header()
        return y_cur

    def draw_table_paginated(all_rows, start_y, col_widths):
        header = all_rows[0]
        rows = all_rows[1:]
        i, y_cur = 0, start_y

        while i < len(rows):
            low, high, fit = 1, len(rows) - i, 0

            while low <= high:
                mid = (low + high) // 2
                trial = [header] + rows[i:i + mid]
                t_try = Table(trial, colWidths=col_widths)
                t_try.setStyle(table_style())
                _, h_try = t_try.wrap(usable_w, 0)

                if y_cur - h_try >= M_B:
                    fit = mid
                    low = mid + 1
                else:
                    high = mid - 1

            if fit == 0:
                pdf_canvas.showPage()
                y_cur = draw_header()
                continue

            chunk = [header] + rows[i:i + fit]
            t = Table(chunk, colWidths=col_widths)
            t.setStyle(table_style())
            w, h = t.wrap(usable_w, 0)
            x = M_L + (usable_w - w) / 2
            t.drawOn(pdf_canvas, x, y_cur - h)
            y_cur -= h + 10
            i += fit

            if i < len(rows) and y_cur < M_B + 60:
                pdf_canvas.showPage()
                y_cur = draw_header()

        return y_cur

    # Draw main table
    y = draw_header()
    y = draw_table_paginated(data, y - 10, col_widths=[100, 80, 80, 120, 80])

    # Draw summaries by supplier
    y -= 30
    y = ensure_space(y, need=50)

    pdf_canvas.setFont("Helvetica-Bold", 14)
    pdf_canvas.drawString(M_L, y, "Summary by Supplier:")
    y -= 20

    for supplier in sorted(by_supplier_mat.keys()):
        y = ensure_space(y, need=30)
        mat_map = by_supplier_mat[supplier]

        pdf_canvas.setFont("Helvetica-Bold", 11)
        pdf_canvas.drawString(M_L, y, f"{supplier}:")
        y -= 16

        pdf_canvas.setFont("Helvetica", 10)
        for mat in sorted([k for k in mat_map.keys() if k != '_subtotal']):
            val = mat_map[mat]["net"]
            pdf_canvas.drawString(M_L + 20, y, f"{mat}: {val:.1f} kg")
            y -= 14

    # Totals
    y -= 20
    y = ensure_space(y, need=40)
    pdf_canvas.setFont("Helvetica-Bold", 12)
    pdf_canvas.drawString(M_L, y, f"Total Gross: {total_gross:.1f} kg")
    y -= 16
    pdf_canvas.drawString(M_L, y, f"Total Net: {total_net:.1f} kg")

    # Finish PDF
    pdf_canvas.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    filename = f"Scales_Daily_{company_slug}_{date_obj.strftime('%Y-%m-%d')}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
def export_monthly_excel(request):
    """
    Excel report for a month (current company).
    Query: ?month=YYYY-MM (defaults to current month)
    """
    company = request.session.get("company_slug")
    month_str = request.GET.get("month") or datetime.now().strftime("%Y-%m")

    try:
        year, month = map(int, month_str.split("-"))
    except ValueError:
        return HttpResponse("Invalid month format. Use YYYY-MM", status=400)

    # Get records
    records = ReceivedMaterial.objects.filter(
        company_slug=company,
        report_day__year=year,
        report_day__month=month
    ).order_by("report_day", "created_at")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    # Header
    ws.append(["Date", "Material", "Gross (kg)", "Net (kg)", "Supplier", "Tag #"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data
    total_gross = 0.0
    total_net = 0.0

    for r in records:
        ws.append([
            r.report_day.strftime("%Y-%m-%d"),
            r.material,
            float(r.gross_kg),
            float(r.net_kg),
            r.supplier,
            r.tag
        ])
        total_gross += float(r.gross_kg)
        total_net += float(r.net_kg)

    # Totals
    ws.append([])
    ws.append(["Totals", "", total_gross, total_net, "", ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Save to buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"monthly_report_{company}_{month_str}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response