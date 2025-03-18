from django.core.serializers import serialize
import os
from .models import Client, Deals, Task, PipeLine, CompanyPallets, Company, Contact, Employee, ContactMaterial,ScheduledShipment

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.decorators import action

from .serializers import ClientSerializer,DealSerializer,PipeLineSerializer

from .forms import ContactForm, CompanyForm, ContactMaterialForm, DealForm
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum
from django.http import JsonResponse
from django.http import HttpResponse
from django.db.models.functions import ExtractYear, ExtractMonth
from django.utils.translation import gettext_lazy as _
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseRedirect
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import json
from decimal import Decimal


from datetime import datetime,timezone,timedelta
from django.utils import timezone as t
from io import BytesIO

from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from .google_calendar import get_calendar_events


from .models import Event



def index(request):
    return render(request, 'crm/index.html')


def client_list(request):
    clients = Client.objects.all()

    supplier = clients.filter(client_type='suppliers')
    buyer = clients.filter(client_type='buyers')

    return render(request, 'crm/client_list.html', {
        'clients': clients,
        'suppliers': supplier,
        'buyers': buyer,
    })


def company_list(request):
    companies = Company.objects.all()
    return render(request, 'crm/company_ main.html', {'companies': companies})


def add_company(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            company = Company.objects.create(name=name)
            # Перенаправляем пользователя на страницу добавления контакта для этой компании
            return redirect('add_contact', company_id=company.id)

    return render(request, 'crm/add_company.html')

# Редактирование компании

def company_detail(request, company_id):
    # Получаем объект компании по id
    company = get_object_or_404(Company, id=company_id)

    # Получаем все контакты этой компании
    contacts = company.contacts.all()

    return render(request, 'crm/company_detail.html', {
        'company': company,
        'contacts': contacts,
    })

def edit_company(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    if request.method == 'POST':
        form = CompanyForm(request.POST, instance=company)
        if form.is_valid():
            form.save()
            return redirect('company_detail', company_id=company.id)
    else:
        form = CompanyForm(instance=company)
    return render(request, 'crm/edit_company.html', {'form': form, 'company': company})

def delete_company(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    if request.method == 'POST':
        company.delete()
        return redirect('company_list')  # После удаления перенаправляем на список компаний
    return render(request, 'crm/delete_company.html', {'company': company})


# Редактирование контакта
def edit_contact(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()
            return redirect('company_list')  # Перенаправить на список компаний
    else:
        form = ContactForm(instance=contact)

    return render(request, 'crm/edit_contact.html', {'form': form, 'contact': contact})


# Удаление контакта
def delete_contact(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    # Удаляем контакт
    contact.delete()

    # Перенаправляем обратно на список компаний
    return redirect('Contacts')  # Указываем имя маршрута для списка компаний

def view_contact(request, id):
    contact = get_object_or_404(Contact, id=id)

    if request.method == "POST":
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()
            return redirect('view_contact', id = contact.id)
    else:
        form=ContactForm(instance=contact)


    return render(request, 'crm/view_contact.html', {'contact': contact, 'form': form})


def manage_employees(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    contacts = company.contacts.all()

    if request.method == 'POST':
        contact_id = request.POST.get('contact_id')
        contact = get_object_or_404(Contact, id=contact_id)

        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        position = request.POST.get('position')

        Employee.objects.create(
            contact=contact,
            name=name,
            email=email,
            phone=phone,
            position=position
        )

        return redirect('manage_employees', company_id=company.id)

    return render(request, 'crm/manage_employees.html', {'company': company, 'contacts': contacts})


def get_employees(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    employees = Employee.objects.filter(contact__company=company)

    # Возвращаем данные сотрудников в формате JSON
    employees_data = [
        {
            'name': employee.name,
            'email': employee.email,
            'phone': employee.phone,
            'position': employee.position,
        }
        for employee in employees
    ]

    return JsonResponse({'employees': employees_data})


def contacts_view(request):
    companies = Company.objects.prefetch_related('contacts').all()
    return render(request, 'crm/contacts_list.html', {'companies': companies})


def add_employee(request, contact_id):
    # Получаем контакт по id
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        # Получаем данные из формы
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        position = request.POST.get('position')

        # Создаем нового сотрудника для этого контакта
        Employee.objects.create(
            contact=contact,
            name=name,
            email=email,
            phone=phone,
            position=position
        )

        # Перенаправляем на страницу контактов после добавления
        return redirect('contacts_list')

    return render(request, 'add_employee.html', {'contact': contact})

def delete_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        employee.delete()
        return redirect('contacts')


def load_employees(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)
    employees = contact.employees.all()
    return render(request, 'employees.html', {'contact': contact, 'employees': employees})


def add_contact(request, company_id):
    company = get_object_or_404(Company, id=company_id)

    if request.method == 'POST':
        # Получаем данные из формы
        address = request.POST.get('address')
        company_type = request.POST.get('company_type')

        # Создаем новый контакт с адресом и типом компании
        Contact.objects.create(
            company=company,
            address=address,
            company_type=company_type
        )

        # Перенаправляем на страницу компании
        return redirect('company_detail', company_id=company.id)

    return render(request, 'crm/add_contact.html', {'company': company})

def company_main(request):
    companies = Company.objects.all()
    return render(request, 'crm/company_main.html', {'companies': companies})

# DEALS
def deal_list(request):
    # Получаем текущий месяц и год
    print("DEBUG: datetime is", datetime)
    today_date = datetime.today()  # ✅ Теперь работает

    current_month = today_date.month
    current_year = today_date.year

    # Получаем базовый набор данных
    deals = Deals.objects.all()

    # Фильтруем компании по типу через связанные контакты
    suppliers = Company.objects.filter(contacts__company_type="suppliers").distinct()  # Только поставщики
    buyers = Company.objects.filter(contacts__company_type="buyers").distinct()  # Только покупатели

    # Получаем параметры фильтра из запроса
    month = request.GET.get('month', str(current_month).zfill(2))  # Текущий месяц по умолчанию
    year = request.GET.get('year', str(current_year))  # Текущий год по умолчанию

    # Применяем фильтры только если месяц и год указаны
    if month and year:
        deals = deals.filter(date__month=int(month), date__year=int(year))
    elif month:  # Если указан только месяц
        deals = deals.filter(date__month=int(month))
    elif year:  # Если указан только год
        deals = deals.filter(date__year=int(year))

    # Подсчёт итогов по отфильтрованным сделкам
    totals = deals.aggregate(
        total_income_loss=Sum('total_income_loss'),
        total_amount=Sum('total_amount'),
    )

    # Получаем доступные года из базы данных
    years = Deals.objects.annotate(year=ExtractYear('date')).values_list('year', flat=True).distinct()

    # Список месяцев для фильтрации (с добавлением пустого значения для "Все")
    months = range(1, 13)  # Месяцы с 1 по 12

    month_names = {
        '01': _('January'),
        '02': _('February'),
        '03': _('March'),
        '04': _('April'),
        '05': _('May'),
        '06': _('June'),
        '07': _('July'),
        '08': _('August'),
        '09': _('September'),
        '10': _('October'),
        '11': _('November'),
        '12': _('December'),
    }

    form = DealForm()

    # Контекст для шаблона
    context = {
        'deals': deals,
        'suppliers': suppliers,  # Только поставщики
        'buyers': buyers,  # Только покупатели, если нужно
        'month': month,
        'year': year,
        'totals': totals,
        'years': sorted(years),  # Сортируем список лет для удобства
        'month_names': month_names,
        'months': months,  # Месяцы для выпадающего списка
        'setting': settings,
        'form': form,
    }

    # Рендерим страницу с переданным контекстом
    return render(request, 'crm/deal_list.html', context)

def export_deals_to_excel(request):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side

    # Получаем текущий месяц и год
    current_date = datetime.now()
    current_month = current_date.month
    current_year = current_date.year

    # Создаем книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"

    # Настройка стилей для заголовков
    header_font = Font(name="Arial", bold=True, color="FFFFFF")  # Белый жирный текст
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Светло-синий фон
    header_alignment = Alignment(horizontal="center", vertical="center")  # Центрирование текста
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Заголовки столбцов
    headers = [
        'Date', 'Supplier', 'Buyer', 'Grade', 'Shipped Qty', 'Pallets',
        'Received Qty', 'Pallets', 'Supplier Price', 'Supplier Paid Amount', 'Buyer Price',
        'Total Amount', 'Transport Cost', 'Houler', 'Income/Loss','Scale Ticket'
    ]
    ws.append(headers)

    # Применяем стили к заголовкам
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Получаем сделки только за текущий месяц
    deals = Deals.objects.select_related('supplier', 'buyer').filter(
        date__year=current_year,
        date__month=current_month
    )

    # Применяем стили к данным
    data_fill_light = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")  # Светлый фон для четных строк
    data_fill_dark = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # Темный фон для нечетных строк
    data_font = Font(name="Arial", size=11)  # Стандартный шрифт для данных

    for row_num, deal in enumerate(deals, start=2):
        formatted_date = deal.date.strftime('%Y-%m') if deal.date else ''
        row = [
            formatted_date,
            deal.supplier.name if deal.supplier else '',
            deal.buyer.name if deal.buyer else '',
            deal.grade,
            deal.shipped_quantity,
            deal.shipped_pallets,
            deal.received_quantity,
            deal.received_pallets,
            deal.supplier_price,
            deal.supplier_total,
            deal.buyer_price,
            deal.total_amount,
            deal.transport_cost,
            deal.transport_company,
            deal.total_income_loss,
            deal.scale_ticket,
        ]

        ws.append(row)

        last_row = len(deals) + 1  # Последняя строка данных

        # Применяем стили к строке данных
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = data_fill_light if row_num % 2 == 0 else data_fill_dark

    # Начало сводной секции справа от основной таблицы
    summary_col_start = len(headers) + 2  # Первая колонка справа от таблицы
    summary_data = [
        ("TOTAL SALE", "=SUM(L2:L{})".format(len(deals) + 1)),  # Общая продажа (Total Amount)
        ("# of pallets", "=SUM(F2:F{})".format(len(deals) + 1)),  # Общее количество паллет
        ("Transportation cost", "=SUM(M2:M{})".format(len(deals) + 1)),  # Транспортные расходы
        ("Suppliers", "=SUM(J2:J{})".format(len(deals) + 1)),  # Итоги для поставщиков
        ("MT OCC11",
         "=SUMPRODUCT((D2:D{0}=\"OCC11\")+(D2:D{0}=\"OCC 11\")+(D2:D{0}=\"OCC 11 Bale String\")+"
         "(D2:D{0}=\"Loose OCC\")+(D2:D{0}=\"Stock Rolls\")+(D2:D{0}=\"Printers Offcuts\"), E2:E{0})".format(
             last_row
         )),
        ("MT Plastic", "=SUMIF(D2:D{}, \"Flexible Plastic\", E2:E{})".format(last_row, last_row)),
        ("MT Mixed-containers", "=SUMIF(D2:D{}, \"Mixed Container\", E2:E{})".format(last_row, last_row)),
        ("INCOME", "=SUM(O2:O{})".format(last_row))
    ]

    # Заполняем сводные данные
    for row_num, (label, value) in enumerate(summary_data, start=2):
        label_cell = ws.cell(row=row_num, column=summary_col_start)
        label_cell.value = label
        label_cell.font = Font(bold=True, name="Arial", size=12)
        label_cell.alignment = Alignment(horizontal="right")

        value_cell = ws.cell(row=row_num, column=summary_col_start + 1)
        value_cell.value = value
        value_cell.border = thin_border
        value_cell.font = Font(name="Arial", size=11)

    # Автоматическая подгонка ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=deals_{current_year}_{current_month}.xlsx'
    wb.save(response)
    return response


def get_deal_details(request, deal_id):
    deal = get_object_or_404(Deals, id=deal_id)

    return JsonResponse({
        'id': deal.id,
        'date': deal.date.strftime('%Y-%m-%d'),
        'supplier': deal.supplier.name if deal.supplier else "",  # ✅ Защита от None
        'buyer': deal.buyer.name if deal.buyer else "",  # ✅ Защита от None
        'grade': deal.grade,
        'shipped_quantity': deal.shipped_quantity,
        'shipped_pallets': deal.shipped_pallets,  # ✅ Добавлено
        'received_quantity': deal.received_quantity,
        'received_pallets': deal.received_pallets,  # ✅ Добавлено
        'supplier_price': deal.supplier_price,
        'buyer_price': deal.buyer_price,
        'total_amount': deal.total_amount,
        'transport_cost': deal.transport_cost,
        'transport_company': deal.transport_company,
        'scale_ticket': deal.scale_ticket,
    })


@csrf_exempt
def edit_deal(request, deal_id):
    deal = get_object_or_404(Deals, id=deal_id)

    if request.method == 'POST':
        data = json.loads(request.body)

        # Получаем объект Company для supplier и buyer по их ID
        supplier_id = data.get('supplier')
        buyer_id = data.get('buyer')

        if supplier_id:
            deal.supplier = get_object_or_404(Company, id=supplier_id)
        if buyer_id:
            deal.buyer = get_object_or_404(Company, id=buyer_id)

        # Преобразуем данные в числовые значения
        deal.received_quantity = Decimal(data.get('received_quantity', deal.received_quantity))
        deal.buyer_price = Decimal(data.get('buyer_price', deal.buyer_price))
        deal.supplier_price = Decimal(data.get('supplier_price', deal.supplier_price))

        # Обновляем остальные поля
        deal.date = data.get('date', deal.date)
        deal.grade = data.get('grade', deal.grade)
        deal.shipped_quantity = data.get('shipped_quantity', deal.shipped_quantity)
        deal.shipped_pallets = data.get('shipped_pallets', deal.shipped_pallets)
        deal.received_pallets = data.get('received_pallets', deal.received_pallets)
        deal.transport_cost = Decimal(data.get('transport_cost', deal.transport_cost))
        deal.transport_company = data.get('transport_company', deal.transport_company)
        deal.scale_ticket = data.get('scale_ticket', deal.scale_ticket)  # ✅ Теперь scale_ticket обновляется

        # Выполняем расчеты для итоговых сумм
        deal.total_amount = deal.received_quantity * deal.buyer_price
        deal.supplier_total = deal.received_quantity * deal.supplier_price
        deal.total_income_loss = deal.total_amount - (deal.supplier_total + deal.transport_cost)

        # Сохраняем обновленную сделку
        deal.save()

        return JsonResponse({
            'status': 'success',
            'deal': {
                'id': deal.id,
                'date': deal.date,
                'supplier': deal.supplier.name,
                'buyer': deal.buyer.name if deal.buyer else '',
                'grade': deal.grade,
                'total_amount': str(deal.total_amount),
                'total_income_loss': str(deal.total_income_loss),
                'scale_ticket': deal.scale_ticket  # ✅ Теперь отправляется обратно в UI
            }
        })

    return JsonResponse({'error': 'Invalid request method'}, status=400)


@csrf_exempt
def delete_deal(request, deal_id):
    if request.method == 'DELETE':
        deal = get_object_or_404(Deals, id=deal_id)
        deal.delete()
        return JsonResponse({'message': 'Deal deleted successfully!'})
    return HttpResponse(status=405)  # Method not allowed


class ClientCreateAPIView(APIView):
    def post(self, request):
        serializer = ClientSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer


class DealCreateAPIView(APIView):
    def post(self, request):
        serializer = DealSerializer(data=request.data)
        if serializer.is_valid():
            scale_ticket = request.data.get("scale_ticket")  # ✅ Берем scale_ticket

            # ✅ Создаем объект, передавая scale_ticket
            deal = serializer.save(scale_ticket=scale_ticket)

            return Response(DealSerializer(deal).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DealViewSet(viewsets.ModelViewSet):
    queryset = Deals.objects.all()
    serializer_class = DealSerializer


def get_licence_plates(request):
    plates = [
        'SY1341',
        'WB3291',
        '153'
    ]
    return JsonResponse({"plates": plates})

def sales_analytics(request):
    from django.db.models import Sum, Q
    from django.db.models.functions import ExtractMonth, ExtractYear
    from django.http import JsonResponse, HttpResponseRedirect
    from datetime import datetime

    # Получаем текущий месяц и год
    now = datetime.now()
    current_month = now.month
    current_year = now.year

    # Получаем параметры фильтра из запроса
    month = request.GET.get('month', str(current_month).zfill(2))  # По умолчанию текущий месяц
    year = request.GET.get('year', str(current_year))               # По умолчанию текущий год

    # --- Данные для графика по месяцам (фильтрация только по году) ---
    deals_year = Deals.objects.filter(date__year=int(year))
    monthly_data = deals_year.values(month=ExtractMonth('date')).annotate(
        total_income=Sum('total_income_loss'),
        total_pallets=Sum('shipped_pallets'),
        transport_cost=Sum('transport_cost'),
        supplier_total=Sum('supplier_total'),
        total_tonnage=Sum('received_quantity'),
        occ11_tonnage=Sum('received_quantity', filter=Q(grade="OCC11") | Q(grade="OCC 11") | Q(grade="Loose OCC") | Q(
            grade="OCC 11 Bale String") | Q(
            grade="Printers Offcuts") | Q(grade="Stock Rolls")),

        plastic_tonnage=Sum('received_quantity', filter=Q(grade="Flexible Plastic")),
            mixed_tonnage=Sum('received_quantity', filter=Q(grade="Mixed Container")),
            total_sales=Sum('total_amount')
        ).order_by('month')

    chart_data = {
        'months': [entry['month'] for entry in monthly_data],
        'profit': [float(entry["total_income"] or 0) for entry in monthly_data],
        'pallets': [float(entry['total_pallets'] or 0) for entry in monthly_data],
        'hauler': [float(entry['transport_cost'] or 0) for entry in monthly_data],
        'suppliers': [float(entry["supplier_total"] or 0) for entry in monthly_data],
        'total_tonnage': [float(entry["total_tonnage"] or 0) for entry in monthly_data],
        'occ11_tonnage': [float(entry["occ11_tonnage"] or 0) for entry in monthly_data],
        'plastic_tonnage': [float(entry["plastic_tonnage"] or 0) for entry in monthly_data],
        'mixed_tonnage': [float(entry["mixed_tonnage"] or 0) for entry in monthly_data],
        'sales': [float(entry["total_sales"] or 0) for entry in monthly_data]
    }

    # Если запрос AJAX, возвращаем данные графика в JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse(chart_data)

    # --- Данные для остальных статистик (фильтрация по месяцу и году) ---
    deals_details = Deals.objects.all()
    if month and year:
        deals_details = deals_details.filter(date__month=int(month), date__year=int(year))
    elif month:
        deals_details = deals_details.filter(date__month=int(month))
    elif year:
        deals_details = deals_details.filter(date__year=int(year))

    suppliers_income = deals_details.values('supplier').annotate(total_income_loss=Sum('total_income_loss'))
    suppliers_income_dict = {
        contact.company.name: float(entry['total_income_loss'] or 0)
        for entry in suppliers_income
        for contact in Contact.objects.filter(company__id=entry['supplier'])
    }

    occ11_filter = Q(grade__iexact="OCC11") | Q(grade__iexact="OCC 11") | Q(grade__iexact="Loose OCC") | \
                   Q(grade__iexact="OCC 11 Bale String") | Q(grade__iexact="Printers Offcuts") | Q(
        grade__iexact="Stock Rolls")

    total_deals = deals_details.count()
    total_sale = deals_details.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_pallets = deals_details.aggregate(Sum('shipped_pallets'))['shipped_pallets__sum'] or 0
    transportation_fee = deals_details.aggregate(Sum('transport_cost'))['transport_cost__sum'] or 0
    suppliers_total = deals_details.aggregate(Sum('supplier_total'))['supplier_total__sum'] or 0
    mt_occ11 = deals_details.filter(occ11_filter).aggregate(Sum('received_quantity'))['received_quantity__sum'] or 0
    mt_plastic = deals_details.filter(grade="Flexible Plastic").aggregate(Sum('received_quantity'))['received_quantity__sum'] or 0
    mt_mixed_containers = deals_details.filter(grade="Mixed Container").aggregate(Sum('received_quantity'))['received_quantity__sum'] or 0
    income = deals_details.aggregate(Sum('total_income_loss'))['total_income_loss__sum'] or 0

    company_pallets = CompanyPallets.objects.select_related('company_name')

    # Обработка формы обновления паллет
    if request.method == 'POST' and 'update_pallets' in request.POST:
        for pallet in company_pallets:
            new_pallet_count = request.POST.get(f"pallets_{pallet.id}")
            new_cages_count = request.POST.get(f"cages_{pallet.id}")  # Получаем данные клеток

            if new_pallet_count is not None:
                pallet.pallets_count = int(new_pallet_count)
            if new_cages_count is not None:
                pallet.cages_count = int(new_cages_count)

            pallet.save()
        return HttpResponseRedirect(request.path)

    years = Deals.objects.annotate(year=ExtractYear('date')).values_list('year', flat=True).distinct()
    months = range(1, 13)

    # Объединяем все данные в один context
    context = {
        'chart_data': chart_data,
        'year': year,
        'suppliers_income': suppliers_income_dict,
        'total_deals': total_deals,
        'total_sale': total_sale,
        'total_pallets': total_pallets,
        'transportation_fee': transportation_fee,
        'suppliers_total': suppliers_total,
        'mt_occ11': mt_occ11,
        'mt_plastic': mt_plastic,
        'mt_mixed_containers': mt_mixed_containers,
        'income': income,
        'company_pallets': company_pallets,
        'month': month,
        'years': sorted(years),
        'months': months,
    }

    return render(request, 'crm/sales_analytics.html', context)


    # Получаем доступные года и месяцы
    years = Deals.objects.annotate(year=ExtractYear('date')).values_list('year', flat=True).distinct()
    months = range(1, 13)

    context = {
        'suppliers_income': suppliers_income_dict,
        'total_deals': total_deals,
        'total_sale': total_sale,
        'total_pallets': total_pallets,
        'transportation_fee': transportation_fee,
        'suppliers_total': suppliers_total,
        'mt_occ11': mt_occ11,
        'mt_plastic': mt_plastic,
        'mt_mixed_containers': mt_mixed_containers,
        'income': income,
        'company_pallets': company_pallets,
        'month': month,
        'year': year,
        'years': sorted(years),
        'months': months,
    }



def add_contact_material(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        # Создаем экземпляр формы с данными из POST-запроса
        form = ContactMaterialForm(request.POST, request.FILES)

        # Проверяем, что форма валидна
        if form.is_valid():
            # Если форма валидна, сохраняем материал
            contact_material = form.save(commit=False)
            contact_material.contact = contact  # Связываем материал с контактом
            contact_material.save()

            # После сохранения редиректим на страницу с контактами
            return redirect('view_contact', id = contact.id)

    else:
        # Если метод запроса GET, создаем пустую форму
        form = ContactMaterialForm()

    # Рендерим шаблон с формой и данными о контакте
    return render(request, 'crm/add_contact_material.html', {'form': form, 'contact': contact})


def edit_contact_material(request, pk):
    contact_material = get_object_or_404(ContactMaterial, pk=pk)

    if request.method == 'POST':
        form = ContactMaterialForm(request.POST, instance=contact_material)
        if form.is_valid():
            form.save()
            return redirect('view_contact', id=contact_material.contact.id)
    else:
        form = ContactMaterialForm(instance=contact_material)

    return render(request, 'crm/edit_contact_material.html', {'form': form, 'contact_material': contact_material})


#REPORTS

def report_list(request):
    return render(request, 'crm/report_list.html')


def company_report(request):
    # Получаем текущий месяц и год
    now = datetime.now()
    current_month = now.month
    current_year = now.year

    # Получаем параметры фильтра из запроса
    selected_company_id = request.GET.get('company', '')
    month = request.GET.get('month', '') or str(current_month).zfill(2)
    year = request.GET.get('year', '') or str(current_year)

    transport_companies = Company.objects.filter(contacts__company_type='hauler')

    # Фильтрация сделок
    deals = Deals.objects.all()

    if selected_company_id:
        # Фильтруем по поставщикам и покупателям
        deals = deals.filter(Q(supplier__id=int(selected_company_id)) |
                             Q(buyer__id=int(selected_company_id)) |
                             Q(transport_company__in=transport_companies)
        )

    if month and year:
        deals = deals.filter(date__month=int(month), date__year=int(year))
    elif month:  # Если указан только месяц
        deals = deals.filter(date__month=int(month))
    elif year:  # Если указан только год
        deals = deals.filter(date__year=int(year))

    # Итоги
    total_transport_cost = deals.aggregate(Sum('transport_cost'))['transport_cost__sum'] or 0
    total_supplier_paid = deals.aggregate(Sum('supplier_total'))['supplier_total__sum'] or 0
    total_amount_buyer = deals.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_buyer_paid = deals.aggregate(Sum('total_income_loss'))[
                           'total_income_loss__sum'] or 0  # Итог для покупателя (например, прибыль или убыток)

    # Список компаний для выбора в фильтре
    companies = Company.objects.all()

    # Получаем доступные года из базы данных для фильтрации
    years = Deals.objects.annotate(year=ExtractYear('date')).values_list('year', flat=True).distinct()

    # Список месяцев для фильтрации
    months = range(1, 13)  # Месяцы с 1 по 12

    # Контекст для шаблона
    context = {
        'deals': deals,
        'total_transport_cost': total_transport_cost,
        'total_supplier_paid': total_supplier_paid,
        'total_amount_buyer': total_amount_buyer,
        'total_buyer_paid': total_buyer_paid,  # Передаем total_buyer_paid в шаблон
        'companies': companies,
        'selected_company_id': int(selected_company_id) if selected_company_id.isdigit() else None,
        'month': month,
        'year': year,
        'years': sorted(years),  # Сортируем список лет для удобства
        'months': months,  # Месяцы для выпадающего списка
    }
    return render(request, 'crm/company_report.html', context)

def export_company_report_pdf(request):
    # Получение фильтров из запроса
    selected_company_id = request.GET.get('company', '')
    month = request.GET.get('month', '')
    year = request.GET.get('year', '')

    # Текущая дата для отчета
    now = datetime.now()

    # Получение всех сделок
    deals = Deals.objects.all()
    if selected_company_id:
        deals = deals.filter(supplier__id=selected_company_id) | \
                deals.filter(buyer__id=selected_company_id) | \
                deals.filter(transport_company__id=selected_company_id)

    # Фильтр по месяцу и году
    if month:
        deals = deals.filter(date__month=int(month))
    if year:
        deals = deals.filter(date__year=int(year))

    # Логика для определения типа компании
    company = None
    company_type = None
    total_field = None  # Поле для итогов в зависимости от компании

    if selected_company_id:
        company = Company.objects.filter(id=selected_company_id).first()
        if company:
            if deals.filter(supplier=company).exists():
                company_type = "supplier"
                deals = deals.filter(supplier=company)
                total_field = "supplier_total"
            elif deals.filter(buyer=company).exists():
                company_type = "buyer"
                deals = deals.filter(buyer=company)
                total_field = "total_amount"
            elif deals.filter(transport_company=company).exists():
                company_type = "transport"
                deals = deals.filter(transport_company=company)
                total_field = "transport_cost"

    # Данные для таблицы
    data = [["Date", "Customer", "Grade", "Net", "Price", "Amount", "Total"]]
    for deal in deals:
        total_value = getattr(deal, total_field, 0)
        data.append([
            deal.date.strftime("%Y-%m-%d"),
            deal.customer.name,
            deal.grade,
            deal.received_quantity,
            deal.price,
            deal.amount,
            total_value
        ])

    # Итоги
    total_net = sum([deal.received_quantity for deal in deals])
    total_amount = sum([getattr(deal, total_field, 0) for deal in deals])

    # Создание PDF
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Лого компании (если есть)
    logo_path = "path/to/your/logo.png"  # Укажите путь к вашему логотипу
    pdf.drawImage(logo_path, x=30, y=height - 80, width=100, height=50)

    # Заголовок
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(140, height - 50, "Company Report")
    pdf.setFont("Helvetica", 10)
    pdf.drawString(30, height - 100, f"Date: {now.strftime('%Y-%m-%d')}")

    # Подзаголовок
    if company:
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(30, height - 120, f"Company: {company.name}")
        pdf.drawString(30, height - 140, f"Type: {company_type.capitalize()}")
    else:
        pdf.drawString(30, height - 120, "All Companies")

    # Таблица
    table = Table(data, colWidths=[80, 80, 80, 50, 50, 50, 50])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN', (3, 1), (6, -1), 'RIGHT'),  # Для чисел
    ]))

    # Установим положение таблицы на странице
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, 30, height - 300)

    # Итоговые данные
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(30, height - 320, f"Total Net: {total_net}")
    pdf.drawString(30, height - 340, f"Total {company_type.capitalize()}: {total_amount}")

    # Завершаем PDF
    pdf.save()
    buffer.seek(0)

    # Возвращаем PDF как ответ
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="company_report_{company_type}.pdf"'
    return response


def get_deal_by_ticket(request):
    ticket_number = request.GET.get('ticket_number', None)
    print(f"Получен запрос на поиск Scale Ticket: {ticket_number}")

    if not ticket_number:
        return JsonResponse({'success': False, 'error': 'No ticket number provided'})

    try:
        deal = Deals.objects.filter(scale_ticket=ticket_number).first()
        if not deal:
            print("Ошибка: Сделка не найдена")
            return JsonResponse({'success': False, 'error': 'No deal found for this ticket number'}, status=404)

        print(f"Найдена сделка: {deal}")

        return JsonResponse({
            'success': True,
            'deal': {
                'id': deal.id,
                'date': deal.date.strftime('%Y-%m-%d'),
                'received_quantity': float(deal.received_quantity),
                'received_pallets': deal.received_pallets,
                'supplier_name': deal.supplier.name if deal.supplier else "",
                'grade': deal.grade,
            }
        })
    except Exception as e:
        import traceback
        print("Ошибка сервера:", traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def export_scale_ticket_pdf(request):
    ticket_number = request.GET.get('ticket_number', None)

    if not ticket_number:
        return HttpResponse("⚠️ No ticket number provided.", status=400)

    # Получаем **все** сделки с таким scale_ticket
    deals = Deals.objects.filter(scale_ticket=ticket_number)

    if not deals.exists():
        return HttpResponse("⚠️ No deals found for this ticket number.", status=404)

    first_deal = deals.first()

    # 📌 Считаем общий вес материалов и паллет
    total_material_weight = sum(deal.received_quantity * 1000 for deal in deals)  # В кг
    total_pallets_weight = sum(deal.received_pallets * 15 for deal in deals if deal.received_pallets)

    # 📌 Получаем данные из формы (если переданы)
    licence_plate = request.GET.get('licence_plate', "N/A")
    tare_weight = float(request.GET.get('tare_weight', 5170))  # 🚛 Базовый вес
    net_weight = float(total_material_weight) + float(total_pallets_weight)  # 📦 Итоговый net_weight
    gross_weight = float(tare_weight) + float(net_weight)  # 📌 Gross = Tare + Net

    # 📌 Форматируем числа
    gross_weight_str = f"{gross_weight:.1f} KG"
    tare_weight_str = f"{tare_weight:.1f} KG"
    net_weight_str = f"{net_weight:.1f} KG"

    # 🕒 Время (из формы или по умолчанию)
    deal_time = request.GET.get('time', "N/A")

    # 🖨 Лог для отладки весов
    print(f"📊 Scale Ticket #{ticket_number} | Gross: {gross_weight}, Tare: {tare_weight}, Net: {net_weight}")

    # Создаём PDF
    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # 🏢 Логотип
    logo_path = os.path.join(settings.BASE_DIR, 'crm', 'static', 'crm', 'images', 'company_logo.png')

    if os.path.exists(logo_path):
        pdf.drawImage(ImageReader(logo_path), 40, height - 80, width=70, height=50, mask='auto')
        print(f"✅ Логотип найден: {logo_path}")
    else:
        print(f"🚨 Логотип НЕ найден: {logo_path}")

    # 📌 Название компании
    pdf.setFont("Helvetica-Bold", 12)
    pdf.setFillColor(colors.darkblue)
    pdf.drawString(130, height - 45, "Local to Global Recycling Inc.")

    # 📍 Адрес
    pdf.setFont("Helvetica", 8)
    pdf.setFillColor(colors.black)
    pdf.drawString(130, height - 55, "19090 Lougheed Hwy.")
    pdf.drawString(130, height - 65, "Pitt Meadows, BC V3Y 2M6")

    # 🏷 Заголовок Scale Ticket
    pdf.setFont("Helvetica", 12)
    pdf.drawString(80, height - 110, f"Scale Ticket #: {ticket_number}")

    # 📆 Дата и время
    pdf.setFont("Helvetica", 10)
    pdf.drawString(80, height - 130, f"Date: {first_deal.date.strftime('%Y-%m-%d')}")
    pdf.drawString(80, height - 150, f"Time: {deal_time}")
    pdf.drawString(80, height - 170, f"Customer:")

    # 👤 Customer details (с переносами строк)
    customer_details = [first_deal.supplier.name] if first_deal.supplier else ["Unknown"]
    y_position = height - 190  # Опускаем ниже заголовка "Customer:"
    for line in customer_details:
        pdf.drawString(85, y_position, line.strip())
        y_position -= 15  # Отступ вниз

    # 📋 Данные справа
    pdf.drawString(350, height - 110, f"Licence: {licence_plate}")
    pdf.drawString(350, height - 130, f"Gross: {gross_weight_str}")
    pdf.drawString(350, height - 150, f"Tare: {tare_weight_str}")
    pdf.drawString(350, height - 170, f"Net: {net_weight_str}")
    pdf.drawString(350, height - 190, f"Pallets #: {first_deal.received_pallets}")

    # 📌 Позиция таблицы (ниже)
    y_position = height - 320

    # 📊 Данные таблицы
    data = [['MATERIAL', 'WEIGHT (KG)', 'PRICE ($/KG)', 'AMOUNT']]
    total_amount = 0

    for deal in deals:
        received_kg = deal.received_quantity * 1000
        sup_price = deal.supplier_price / 1000  # ✅ Цена за кг
        amount = received_kg * sup_price
        total_amount += amount

        data.append([deal.grade, f"{received_kg:.1f}", f"${sup_price:.2f}", f"${amount:.2f}"])

    # 🏋 Добавляем вес паллет, если есть
    if total_pallets_weight > 0:
        data.append(['Pallets', f"{total_pallets_weight:.1f}", '', ''])

    # 📑 Создаем таблицу
    table = Table(data, colWidths=[200, 100, 100, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
    ]))

    # 🖨 Вывод таблицы в PDF
    table.wrapOn(pdf, width, height)
    table.drawOn(pdf, 80, y_position)

    # 💰 Итоговая сумма
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(80, y_position - 25, f"Total: ${total_amount:.2f}")

    # 🛟 Сохранение PDF
    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename=\"Ticket # {ticket_number}.pdf\"'
    return response


# Область доступа
SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CREDENTIALS_FILE = os.path.join(BASE_DIR, "c_id.json")
TOKEN_FILE = os.path.join(BASE_DIR, "token.json")

# ID календаря
CALENDAR_ID = "dmitry@wastepaperbrokers.com"

def get_calendar_events():
    """Получает события из Google Calendar."""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        try:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=8081, access_type="offline", prompt="consent")
            with open(TOKEN_FILE, "w") as token:
                token.write(creds.to_json())
        except Exception as e:
            return {"error": f"Ошибка авторизации: {str(e)}"}

    try:
        service = build("calendar", "v3", credentials=creds)
        now = datetime.now(t.utc).replace(microsecond=0).isoformat() + "Z"

        events_result = service.events().list(
            calendarId=CALENDAR_ID,
            timeMin=now,
            maxResults=20,
            singleEvents=True,
            orderBy="startTime",
        ).execute()

        events = events_result.get("items", [])
    except Exception as e:
        return {"error": f"Ошибка загрузки событий: {str(e)}"}

    return events

def task_list(request):
    suppliers = Company.objects.filter(contacts__company_type="suppliers")
    buyers = Company.objects.filter(contacts__company_type="buyers")

    context = {
        'suppliers': suppliers,
        'buyers': buyers,
        'materials_list': settings.MATERIALS_LIST,  # ✅ Передаем MATERIALS_LIST правильно
    }
    return render(request, 'crm/task_list.html', context)


def get_events(request):
    """Возвращает события в формате JSON для FullCalendar"""
    events = Event.objects.all()
    event_list = [
        {
            "id": event.id,
            "title": event.title,
            "start": event.start.strftime("%Y-%m-%dT%H:%M:%S"),
            "end": event.end.strftime("%Y-%m-%dT%H:%M:%S") if event.end else None,
            "allDay": event.all_day,
        }
        for event in events
    ]
    return JsonResponse(event_list, safe=False)

@csrf_exempt
def add_event(request):
    """Добавляет новое событие в календарь"""
    try:
        data = json.loads(request.body)
        title = data.get("title")
        start = data.get("start")

        if not title or not start:
            return JsonResponse({"error": "title и start обязательны"}, status=400)

        # ✅ Преобразуем дату
        start_datetime = datetime.fromisoformat(start)

        # ✅ Проверяем, есть ли уже таймзона, если нет — добавляем
        if start_datetime.tzinfo is None:
            start_datetime = timezone.make_aware(start_datetime)

        # ✅ Создаем событие в БД
        event = Event.objects.create(
            title=title,
            start=start_datetime,
            all_day=False
        )

        return JsonResponse({"status": "success", "event": {
            "id": event.id,
            "title": event.title,
            "start": event.start.isoformat(),
            "allDay": event.all_day
        }})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)




@csrf_exempt
def delete_event(request, event_id):
    """Удаляет событие"""
    if request.method == "DELETE":
        event = Event.objects.filter(id=event_id).first()
        if event:
            event.delete()
            return JsonResponse({"status": "deleted"})
        return JsonResponse({"error": "Event not found"}, status=404)

    return JsonResponse({"error": "Invalid request"}, status=400)



def get_grades(request):
    """
    Возвращает список всех материалов (грейдов).
    """
    return JsonResponse(list(settings.MATERIALS_LIST.keys()), safe=False)


@csrf_exempt
def add_scheduled_shipment(request):
    """
    Добавляет новую запланированную отгрузку.
    """
    if request.method == "POST":
        data = json.loads(request.body)
        try:
            supplier = Company.objects.get(id=data["supplier"])
            buyer = Company.objects.get(id=data["buyer"])

            shipment = ScheduledShipment.objects.create(
                supplier=supplier,
                buyer=buyer,
                date=data["datetime"].split("T")[0],
                time=data["datetime"].split("T")[1],
                grade=data["grade"]
            )
            return JsonResponse({"status": "success", "shipment_id": shipment.id})
        except Company.DoesNotExist:
            return JsonResponse({"error": "Supplier or Buyer not found"}, status=400)

    return JsonResponse({"error": "Invalid request"}, status=400)

def get_scheduled_shipments(request):
    """
    Возвращает список всех запланированных отгрузок.
    """
    shipments = ScheduledShipment.objects.all().order_by("date")
    shipment_list = [
        {
            "id": shipment.id,
            "date": shipment.date.strftime("%Y-%m-%d"),
            "time": shipment.time.strftime("%H:%M"),
            "supplier": shipment.supplier.name,
            "buyer": shipment.buyer.name,
            "grade": shipment.grade
        }
        for shipment in shipments
    ]
    return JsonResponse(shipment_list, safe=False)

@csrf_exempt
def delete_scheduled_shipment(request, shipment_id):
    """
    Удаляет запланированную отгрузку.
    """
    if request.method == "DELETE":
        try:
            shipment = ScheduledShipment.objects.get(id=shipment_id)
            shipment.delete()
            return JsonResponse({"status": "deleted"})
        except ScheduledShipment.DoesNotExist:
            return JsonResponse({"error": "Shipment not found"}, status=404)

    return JsonResponse({"error": "Invalid request"}, status=400)


def pipeline(request):
    pipeline = PipeLine.objects.all()
    return render(request, 'crm/pipeline_list.html', {'pipeline': pipeline})

class PipelineViewSet(viewsets.ModelViewSet):
    queryset = PipeLine.objects.all()
    serializer_class = PipeLineSerializer

    @action(detail=True, methods=['post'])
    def move(self, request, pk=None):
        """Обновление стадии компании"""
        pipeline = self.get_object()
        new_stage = request.data.get('stage')

        if new_stage not in dict(PipeLine.STAGES):
            return Response({'error': 'Invalid stage'}, status=status.HTTP_400_BAD_REQUEST)

        pipeline.stage = new_stage
        pipeline.save()
        return Response({'status': 'stage updated'})

def pipeline_list(request):
    return JsonResponse({"message": "This is the pipeline list."})