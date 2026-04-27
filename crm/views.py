from django.core.serializers import serialize
import os
from .models import (Client, Deals, Task, PipeLine, CompanyPallets, Company, Contact, Employee, ContactMaterial,
                     ScheduledShipment,SCaleTicketStatus,TruckProfile,EmailRecipientPreference,ExportShipment, ExportLane,
                     VesselSchedule, Deals,ExportShipment, ExportShipmentDocument,ExportDocument)


from datetime import datetime, date, time, timedelta
from decimal import Decimal
from django.utils import timezone
from django.db.models import Sum, Count, F, DecimalField, ExpressionWrapper,Value
from django.db.models.functions import Coalesce

from django.db import transaction

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.decorators import action
from .forms import TaskForm
from .serializers import ClientSerializer,DealSerializer,PipeLineSerializer

from django.http import HttpResponseBadRequest, Http404
from urllib.parse import unquote
from .forms import ContactForm, CompanyForm, ContactMaterialForm, DealForm
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum
from django.http import JsonResponse
from django.http import HttpResponse
from django.db.models.functions import ExtractYear, ExtractMonth
from django.utils.translation import gettext_lazy as _
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponseRedirect
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.views.decorators.http import require_POST
import json

from pathlib import Path
from datetime import datetime,timezone,timedelta
from django.utils import timezone as t
from io import BytesIO

from .supplier_shipment_report_archive import generate_supplier_shipment_reports_for_month

from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from .google_calendar import get_calendar_events
import glob

from django.utils.timezone import make_aware
from django.db.models import Sum, Count, F
from urllib.parse import unquote

from .models import Event
from django.test import RequestFactory

import re

def sanitize_filename(name):
    name = name.strip()
    name = name.replace(' ', '_')
    return re.sub(r'[<>:"/\\|?*]', '_', name)



from django.conf import settings
print("📦 Current database path:", settings.DATABASES['default']['NAME'])


def index(request):
    return render(request, 'crm/index.html')


def client_list(request):
    clients = Client.objects.all()

    supplier = clients.filter(client_type='suppliers')
    buyer = clients.filter(client_type='buyers')
    hauler = clients.filter(client_type='hauler')
    return render(request, 'crm/client_list.html', {
        'clients': clients,
        'suppliers': supplier,
        'buyers': buyer,
        'hauler': hauler,
    })


def company_list(request):
    companies = Company.objects.all()
    return render(request, 'crm/company_ main.html', {'companies': companies})


def add_company(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            company = Company.objects.create(name=name)
            # Перенаправляем пользователя на страницу добавления контакта для этой компании
            return redirect('add_contact', company_id=company.id)

    return render(request, 'crm/add_company.html')

# Редактирование компании

def company_detail(request, company_id):
    # Получаем объект компании по id
    company = get_object_or_404(Company, id=company_id)

    # Получаем все контакты этой компании
    contacts = company.contacts.all()

    return render(request, 'crm/company_detail.html', {
        'company': company,
        'contacts': contacts,
    })

def edit_company(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    if request.method == 'POST':
        form = CompanyForm(request.POST, instance=company)
        if form.is_valid():
            form.save()
            return redirect('company_detail', company_id=company.id)
    else:
        form = CompanyForm(instance=company)
    return render(request, 'crm/edit_company.html', {'form': form, 'company': company})

def delete_company(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    if request.method == 'POST':
        company.delete()
        return redirect('company_list')  # После удаления перенаправляем на список компаний
    return render(request, 'crm/delete_company.html', {'company': company})


# Редактирование контакта
def edit_contact(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()
            return redirect('company_list')  # Перенаправить на список компаний
    else:
        form = ContactForm(instance=contact)

    return render(request, 'crm/edit_contact.html', {'form': form, 'contact': contact})


# Удаление контакта
def delete_contact(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    # Удаляем контакт
    contact.delete()

    # Перенаправляем обратно на список компаний
    return redirect('Contacts')  # Указываем имя маршрута для списка компаний

def view_contact(request, id):
    contact = get_object_or_404(Contact, id=id)
    pipeline, _ = PipeLine.objects.get_or_create(contact=contact)

    if request.method == "POST":
        # 🎯 Обработка смены стадии
        if "change_stage" in request.POST:
            stage = request.POST.get("stage")
            pipeline.stage = stage
            pipeline.save()

        # 🎯 Обработка формы редактирования контакта
        form = ContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()

            # 🟡 Сохраняем чекбокс
            pickup_requested = request.POST.get("pickup_requested") == "on"
            contact.company.pickup_requested = pickup_requested
            contact.company.save()

        return redirect("view_contact", id=contact.id)

    else:
        form = ContactForm(instance=contact)

    employees = contact.employees.all()
    trucks = TruckProfile.objects.filter(company=contact.company)

    return render(request, 'crm/view_contact.html', {
        'contact': contact,
        'form': form,
        'employees': employees,
        'pipeline': pipeline,
        'company': contact.company,
        'trucks': trucks,
    })


@csrf_exempt
def toggle_pickup(request, id):
    if request.method == "POST":
        import json
        company = get_object_or_404(Company, id=id)
        data = json.loads(request.body)
        company.pickup_requested = data.get('pickup_requested', False)
        company.save()
        return JsonResponse({"status": "ok"})
    return JsonResponse({"status": "error"}, status=400)

def manage_employees(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    contacts = company.contacts.all()

    if request.method == 'POST':
        contact_id = request.POST.get('contact_id')
        contact = get_object_or_404(Contact, id=contact_id)

        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        position = request.POST.get('position')

        Employee.objects.create(
            contact=contact,
            name=name,
            email=email,
            phone=phone,
            position=position
        )

        return redirect('manage_employees', company_id=company.id)

    return render(request, 'crm/manage_employees.html', {'company': company, 'contacts': contacts})

def edit_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)

    if request.method == 'POST':
        employee.name = request.POST.get('name')
        employee.email = request.POST.get('email')
        employee.phone = request.POST.get('phone')
        employee.position = request.POST.get('position')
        employee.save()
        return redirect('view_contact', id=employee.contact.id)

    return render(request, 'crm/edit_employee.html', {'employee': employee})


def get_employees(request, company_id):
    company = get_object_or_404(Company, id=company_id)
    employees = Employee.objects.filter(contact__company=company)

    # Возвращаем данные сотрудников в формате JSON
    employees_data = [
        {
            'name': employee.name,
            'email': employee.email,
            'phone': employee.phone,
            'position': employee.position,
        }
        for employee in employees
    ]

    return JsonResponse({'employees': employees_data})


def contacts_view(request):
    companies = Company.objects.prefetch_related('contacts').all()
    return render(request, 'crm/contacts_list.html', {'companies': companies})


def add_employee(request, contact_id):
    # Получаем контакт по id
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        # Получаем данные из формы
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        position = request.POST.get('position')

        # Создаем нового сотрудника для этого контакта
        Employee.objects.create(
            contact=contact,
            name=name,
            email=email,
            phone=phone,
            position=position
        )

        # Перенаправляем на страницу контактов после добавления
        return redirect('contacts_list')

    return render(request, 'crm/add_employee.html', {'contact': contact})

def delete_employee(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    if request.method == 'POST':
        employee.delete()
        return redirect('contacts_list')


def load_employees(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)
    employees = contact.employees.all()
    return render(request, 'employees.html', {'contact': contact, 'employees': employees})


def add_contact(request, company_id):
    company = get_object_or_404(Company, id=company_id)

    if request.method == 'POST':
        # Получаем данные из формы
        address = request.POST.get('address')
        company_type = request.POST.get('company_type')

        # Создаем новый контакт с адресом и типом компании
        Contact.objects.create(
            company=company,
            address=address,
            company_type=company_type
        )

        # Перенаправляем на страницу компании
        return redirect('company_detail', company_id=company.id)

    return render(request, 'crm/add_contact.html', {'company': company})

def company_main(request):
    companies = Company.objects.all()
    return render(request, 'crm/company_main.html', {'companies': companies})

# DEALS
def deal_list(request):
    # Получаем текущий месяц и год
    print("DEBUG: datetime is", datetime)
    today_date = datetime.today()
    companies = Company.objects.all()
    current_month = today_date.month
    current_year = today_date.year

    # Получаем базовый набор данных
    deals = Deals.objects.all().order_by('date')

    # Фильтруем компании по типу через связанные контакты
    suppliers = Company.objects.filter(contacts__company_type="suppliers").distinct()  # Только поставщики
    buyers = Company.objects.filter(contacts__company_type="buyers").distinct()  # Только покупатели
    hauler = Company.objects.filter(contacts__company_type="hauler").distinct() # Only Haulers

    # Получаем параметры фильтра из запроса
    month = request.GET.get('month', str(current_month).zfill(2))  # Текущий месяц по умолчанию
    year = request.GET.get('year', str(current_year))  # Текущий год по умолчанию

    selected_company_id = request.GET.get('company')

    if selected_company_id:
        deals = deals.filter(
            Q(supplier__id=selected_company_id) |
            Q(buyer__id=selected_company_id) |
            Q(transport_company__id=selected_company_id)
        )

    # Применяем фильтры только если месяц и год указаны
    if month and year:
        deals = deals.filter(date__month=int(month), date__year=int(year))
    elif month:  # Если указан только месяц
        deals = deals.filter(date__month=int(month))
    elif year:  # Если указан только год
        deals = deals.filter(date__year=int(year))

    # Подсчёт итогов по отфильтрованным сделкам
    totals = deals.aggregate(
        total_income_loss=Sum('total_income_loss'),
        total_amount=Sum('total_amount'),
        total_tonnage =Sum('received_quantity')
    )

    per_grade = (
        deals.values('grade')
        .annotate(
            shipped_weight=Coalesce(
                Sum('received_quantity', output_field=DecimalField(max_digits=18, decimal_places=2)),
                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
            ),
            total_cash=Coalesce(
                Sum('total_amount', output_field=DecimalField(max_digits=18, decimal_places=2)),
                Value(0, output_field=DecimalField(max_digits=18, decimal_places=2))
            ),
        )
        .order_by('grade')
    )

    for row in per_grade:
        row["display"] = f"{row['grade']} – {row['shipped_weight']:.2f} MT ({row['total_cash']:.2f} $)"




    # Получаем доступные года из базы данных
    years = Deals.objects.annotate(year=ExtractYear('date')).values_list('year', flat=True).distinct()

    # Список месяцев для фильтрации (с добавлением пустого значения для "Все")
    months = range(1, 13)  # Месяцы с 1 по 12

    month_names = {
        '01': _('January'),
        '02': _('February'),
        '03': _('March'),
        '04': _('April'),
        '05': _('May'),
        '06': _('June'),
        '07': _('July'),
        '08': _('August'),
        '09': _('September'),
        '10': _('October'),
        '11': _('November'),
        '12': _('December'),
    }

    form = DealForm()

    # Контекст для шаблона
    context = {
        'deals': deals,
        'suppliers': suppliers,  # Только поставщики
        'buyers': buyers,  # Только покупатели, если нужно
        'month': month,
        'year': year,
        'totals': totals,
        'years': sorted(years),  # Сортируем список лет для удобства
        'month_names': month_names,
        'months': months,  # Месяцы для выпадающего списка
        'setting': settings,
        'form': form,
        'hauler': hauler,
        'selected_company_id': int(selected_company_id) if selected_company_id else None,
        'companies': companies,
        'per_grade': per_grade,
    }

    # Рендерим страницу с переданным контекстом
    return render(request, 'crm/deal_list.html', context)




def get_price_by_supplier_and_grade(request):
    supplier_id = request.GET.get('supplier_id')
    grade = request.GET.get('grade')

    try:
        contact_material = ContactMaterial.objects.get(contact__company__id=supplier_id, material=grade)
        return JsonResponse({'price': str(contact_material.price)})
    except ContactMaterial.DoesNotExist:
        return JsonResponse({'price': None})


def get_price_by_buyer_and_grade(request):
    buyer_id = request.GET.get('buyer_id')
    grade = request.GET.get('grade')

    try:
        contact_material = ContactMaterial.objects.get(contact__company__id=buyer_id, material=grade)
        return JsonResponse({'price': str(contact_material.price)})
    except ContactMaterial.DoesNotExist:
        return JsonResponse({'price': None})


def export_deals_to_excel(request):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side

    # Получаем текущий месяц и год
    from django.utils.timezone import now

    month = request.GET.get('month')
    year = request.GET.get('year')

    if not month or not year:
        current_date = now()
        month = current_date.month
        year = current_date.year
    else:
        month = int(month)
        year = int(year)

    # Создаем книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"

    # Настройка стилей для заголовков
    header_font = Font(name="Arial", bold=True, color="FFFFFF")  # Белый жирный текст
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Светло-синий фон
    header_alignment = Alignment(horizontal="center", vertical="center")  # Центрирование текста
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Заголовки столбцов
    headers = [
        'Date', 'Supplier', 'Buyer', 'Grade', 'Shipped Qty', 'Pallets',
        'Received Qty', 'Pallets', 'Supplier Price', 'Supplier Paid Amount', 'Buyer Price',
        'Total Amount', 'Transport Cost', 'Hauler', 'Income/Loss','Scale Ticket'
    ]
    ws.append(headers)

    # Применяем стили к заголовкам
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Получаем сделки только за текущий месяц
    deals = Deals.objects.select_related('supplier', 'buyer').filter(
        date__year=year,
        date__month=month
    )

    # Применяем стили к данным
    data_fill_light = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")  # Светлый фон для четных строк
    data_fill_dark = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # Темный фон для нечетных строк
    data_font = Font(name="Arial", size=11)  # Стандартный шрифт для данных

    last_row = len(deals) + 1  # Последняя строка данных

    for row_num, deal in enumerate(deals, start=2):
        formatted_date = deal.date.strftime('%d-%b') if deal.date else ''

        grade_value = deal.grade
        if grade_value == 'Baled Cardboard':
            grade_value = 'OCC 11'


        row = [
            formatted_date,
            deal.supplier.name if deal.supplier else '',
            deal.buyer.name if deal.buyer else '',
            grade_value,
            deal.shipped_quantity,
            deal.shipped_pallets,
            deal.received_quantity,
            deal.received_pallets,
            deal.supplier_price,
            deal.supplier_total,
            deal.buyer_price,
            deal.total_amount,
            deal.transport_cost,
            deal.transport_company.name if deal.transport_company else '',
            deal.total_income_loss,
            deal.scale_ticket,
        ]

        ws.append(row)



        # Применяем стили к строке данных
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            cell.fill = data_fill_light if row_num % 2 == 0 else data_fill_dark

    # Начало сводной секции справа от основной таблицы
    summary_col_start = len(headers) + 2  # Первая колонка справа от таблицы
    summary_data = [
        ("TOTAL SALE", "=SUM(L2:L{})".format(len(deals) + 1)),  # Общая продажа (Total Amount)
        ("# of pallets", "=SUM(F2:F{})".format(len(deals) + 1)),  # Общее количество паллет
        ("Transportation cost", "=SUM(M2:M{})".format(len(deals) + 1)),  # Транспортные расходы
        ("Suppliers", "=SUM(J2:J{})".format(len(deals) + 1)),  # Итоги для поставщиков
        ("MT OCC11",
         "=SUMPRODUCT((D2:D{0}=\"OCC11\")+(D2:D{0}=\"OCC 11\")+(D2:D{0}=\"OCC 11 Bale String\")+"
         "(D2:D{0}=\"Loose OCC\")+(D2:D{0}=\"Stock Rolls\")+(D2:D{0}=\"Printers Offcuts\"), E2:E{0})".format(
             last_row
         )),
        ("MT Plastic", "=SUMIF(D2:D{}, \"Flexible Plastic\", E2:E{})".format(last_row, last_row)),
        ("MT Mixed-containers", "=SUMIF(D2:D{}, \"Mixed Container\", E2:E{})".format(last_row, last_row)),
        ("INCOME", "=SUM(O2:O{})".format(last_row))
    ]

    # Заполняем сводные данные
    for row_num, (label, value) in enumerate(summary_data, start=2):
        label_cell = ws.cell(row=row_num, column=summary_col_start)
        label_cell.value = label
        label_cell.font = Font(bold=True, name="Arial", size=12)
        label_cell.alignment = Alignment(horizontal="right")

        value_cell = ws.cell(row=row_num, column=summary_col_start + 1)
        value_cell.value = value
        value_cell.border = thin_border
        value_cell.font = Font(name="Arial", size=11)

    # Автоматическая подгонка ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename=deals_{year}_{month}.xlsx'
    wb.save(response)
    return response


def get_deal_details(request, deal_id):
    deal = get_object_or_404(Deals, id=deal_id)

    # Безопасные числа
    shipped_qty = deal.shipped_quantity or Decimal("0")       # MT
    received_qty = deal.received_quantity or Decimal("0")     # MT
    shipped_pallets = deal.shipped_pallets or Decimal("0")
    supplier_price = deal.supplier_price or Decimal("0")      # $/MT
    buyer_price = deal.buyer_price or Decimal("0")            # $/MT
    transport_cost = deal.transport_cost or Decimal("0")      # $
    total_amount = deal.total_amount or Decimal("0")          # $ (выручка)
    total_income_loss = deal.total_income_loss or Decimal("0")  # $ (прибыль/убыток)

    # 1) Прибыль уже есть: total_income_loss

    # 2) Прибыль на тонну
    if shipped_qty > 0:
        profit_per_ton = total_income_loss / shipped_qty      # $/MT
    else:
        profit_per_ton = None

    # 3) Транспорт на тонну
    if shipped_qty > 0:
        transport_per_ton = transport_cost / shipped_qty      # $/MT
    else:
        transport_per_ton = None

    # 4) Доля логистики в выручке
    if total_amount > 0:
        transport_share = (transport_cost / total_amount) * Decimal("100")  # %
    else:
        transport_share = None

    # 5) Ценовой спред (buyer - supplier)
    spread_per_ton = buyer_price - supplier_price             # $/MT

    # 6) Отклонение по весу (MT)
    variance_mt = received_qty - shipped_qty                  # MT

    # 7) Средний вес паллеты (в кг)
    if shipped_qty > 0 and shipped_pallets > 0:
        avg_pallet_weight_kg = (shipped_qty * Decimal("1000")) / shipped_pallets
    else:
        avg_pallet_weight_kg = None

    return JsonResponse({
        "id": deal.id,
        "date": deal.date.strftime("%Y-%m-%d"),

        "supplier_id": deal.supplier.id if deal.supplier else None,
        "supplier_name": deal.supplier.name if deal.supplier else "",
        "buyer_id": deal.buyer.id if deal.buyer else None,
        "buyer_name": deal.buyer.name if deal.buyer else "",
        "grade": deal.grade,

        "shipped_quantity": float(shipped_qty),
        "shipped_pallets": float(shipped_pallets),
        "received_quantity": float(received_qty),
        "received_pallets": float(deal.received_pallets or 0),
        "supplier_price": float(supplier_price),
        "buyer_price": float(buyer_price),
        "total_amount": float(total_amount),
        "transport_cost": float(transport_cost),
        "scale_ticket": deal.scale_ticket or "",
        "transport_company_id": deal.transport_company.id if deal.transport_company else None,
        "transport_company_name": deal.transport_company.name if deal.transport_company else "",

        # KPI
        "total_income_loss": float(total_income_loss),
        "profit_per_ton": float(profit_per_ton) if profit_per_ton is not None else None,
        "transport_per_ton": float(transport_per_ton) if transport_per_ton is not None else None,
        "transport_share": float(transport_share) if transport_share is not None else None,
        "spread_per_ton": float(spread_per_ton),
        "variance_mt": float(variance_mt),
        "avg_pallet_weight_kg": float(avg_pallet_weight_kg) if avg_pallet_weight_kg is not None else None,
    })

@csrf_exempt
def edit_deal(request, deal_id):
    deal = get_object_or_404(Deals, id=deal_id)

    if request.method == 'POST':
        data = json.loads(request.body)

        supplier_id = data.get('supplier')
        buyer_id = data.get('buyer')

        if supplier_id:
            deal.supplier = get_object_or_404(Company, id=supplier_id)
        if buyer_id:
            deal.buyer = get_object_or_404(Company, id=buyer_id)

        # Числовые значения как Decimal
        deal.shipped_quantity = Decimal(data.get('shipped_quantity', deal.shipped_quantity))
        deal.received_quantity = Decimal(data.get('received_quantity', deal.received_quantity))
        deal.buyer_price = Decimal(data.get('buyer_price', deal.buyer_price))
        deal.supplier_price = Decimal(data.get('supplier_price', deal.supplier_price))
        deal.shipped_pallets = Decimal(data.get('shipped_pallets', deal.shipped_pallets))
        deal.received_pallets = Decimal(data.get('received_pallets', deal.received_pallets))
        deal.transport_cost = Decimal(data.get('transport_cost', deal.transport_cost))

        # Остальные поля
        deal.date = data.get('date', deal.date)
        deal.grade = data.get('grade', deal.grade)

        transport_company_id = data.get('transport_company')
        if transport_company_id:
            deal.transport_company = get_object_or_404(Company, id=transport_company_id)

        deal.scale_ticket = data.get('scale_ticket', deal.scale_ticket)

        # Итоговые расчёты
        deal.total_amount = deal.shipped_quantity * deal.buyer_price
        deal.supplier_total = deal.received_quantity * deal.supplier_price
        deal.total_income_loss = deal.total_amount - (deal.supplier_total + deal.transport_cost)

        deal.save()

        return JsonResponse({
            'status': 'success',
            'deal': {
                'id': deal.id,
                'date': deal.date,
                'supplier': deal.supplier.name,
                'buyer': deal.buyer.name if deal.buyer else '',
                'grade': deal.grade,
                'total_amount': str(deal.total_amount),
                'total_income_loss': str(deal.total_income_loss),
                'scale_ticket': deal.scale_ticket
            }
        })

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@csrf_exempt
def delete_deal(request, deal_id):
    if request.method == 'DELETE':
        deal = get_object_or_404(Deals, id=deal_id)
        deal.delete()
        return JsonResponse({'message': 'Deal deleted successfully!'})
    return HttpResponse(status=405)  # Method not allowed


class ClientCreateAPIView(APIView):
    def post(self, request):
        serializer = ClientSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer


class DealCreateAPIView(APIView):
    def post(self, request):
        serializer = DealSerializer(data=request.data)
        if serializer.is_valid():
            scale_ticket = request.data.get("scale_ticket")  # ✅ Берем scale_ticket

            # ✅ Создаем объект, передавая scale_ticket
            deal = serializer.save(scale_ticket=scale_ticket)

            return Response(DealSerializer(deal).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DealViewSet(viewsets.ModelViewSet):
    queryset = Deals.objects.all()
    serializer_class = DealSerializer


def get_licence_plates(request):
    plates = [
        'SY1341',
        'WB3291',
        '153'
    ]
    return JsonResponse({"plates": plates})


SCALE_TICKET_COUNTER_FILE = os.path.join(settings.BASE_DIR, 'scale_ticket_counter.json')



@csrf_exempt
def get_scale_ticket_counters(request):


    if request.method == 'GET':
        if not os.path.exists(SCALE_TICKET_COUNTER_FILE):
            with open(SCALE_TICKET_COUNTER_FILE, 'w') as f:
                json.dump({"scale_ticket": 108346}, f)


        with open(SCALE_TICKET_COUNTER_FILE, 'r') as f:
            data = json.load(f)
        return JsonResponse(data)

    return JsonResponse({"error": "Invalid request"}, status=400)


@csrf_exempt
def increment_scale_ticket_counters(request):
    if request.method == 'POST':
        # ✅ Если файл не существует, создаём его с начальными значениями
        if not os.path.exists(SCALE_TICKET_COUNTER_FILE):
            with open(SCALE_TICKET_COUNTER_FILE, 'w') as f:
                json.dump({"bol": 1000, "load": 2000}, f)

        # ✅ Читаем и обновляем
        with open(SCALE_TICKET_COUNTER_FILE, 'r+') as f:
            data = json.load(f)
            data["scale_ticket"] += 1
            f.seek(0)
            json.dump(data, f, indent=2)
            f.truncate()

        return JsonResponse({"status": "updated", "scale_ticket": data["scale_ticket"]})
    else:
        return JsonResponse({"error": "Invalid request"}, status=400)




def sales_analytics(request):
    from django.db.models import Sum, Q
    from django.db.models.functions import ExtractMonth, ExtractYear
    from django.http import JsonResponse, HttpResponseRedirect
    from datetime import datetime

    # Получаем текущий месяц и год
    now = datetime.now()
    current_month = now.month
    current_year = now.year

    # Получаем параметры фильтра из запроса
    month = request.GET.get('month', str(current_month).zfill(2))  # По умолчанию текущий месяц
    year = request.GET.get('year', str(current_year))               # По умолчанию текущий год

    # --- Данные для графика по месяцам (фильтрация только по году) ---
    deals_year = Deals.objects.filter(date__year=int(year))
    monthly_data = deals_year.values(month=ExtractMonth('date')).annotate(
        total_income=Sum('total_income_loss'),
        total_pallets=Sum('shipped_pallets'),
        transport_cost=Sum('transport_cost'),
        supplier_total=Sum('supplier_total'),
        total_tonnage=Sum('shipped_quantity'),
        occ11_tonnage=Sum(
            'shipped_quantity',
            filter=(
                    Q(grade="OCC11") |
                    Q(grade="OCC 11") |
                    Q(grade="OCC 11 Bale String") |
                    Q(grade="Loose OCC") |
                    Q(grade="Printers Offcuts") |
                    Q(grade="Stock Rolls") |
                    Q(grade="Cardboard Stock Lots") |
                    Q(grade="DLK") |
                    Q(grade='Baled Cardboard')
            )
        ),

        plastic_tonnage=Sum('received_quantity', filter=Q(grade="Flexible Plastic")),
            mixed_tonnage=Sum('received_quantity', filter=Q(grade="Mixed Container")),
            total_sales=Sum('total_amount')
        ).order_by('month')

    chart_data = {
        'months': [entry['month'] for entry in monthly_data],
        'profit': [float(entry["total_income"] or 0) for entry in monthly_data],
        'pallets': [float(entry['total_pallets'] or 0) for entry in monthly_data],
        'hauler': [float(entry['transport_cost'] or 0) for entry in monthly_data],
        'suppliers': [float(entry["supplier_total"] or 0) for entry in monthly_data],
        'total_tonnage': [float(entry["total_tonnage"] or 0) for entry in monthly_data],
        'occ11_tonnage': [float(entry["occ11_tonnage"] or 0) for entry in monthly_data],
        'plastic_tonnage': [float(entry["plastic_tonnage"] or 0) for entry in monthly_data],
        'mixed_tonnage': [float(entry["mixed_tonnage"] or 0) for entry in monthly_data],
        'sales': [float(entry["total_sales"] or 0) for entry in monthly_data]
    }

    # Если запрос AJAX, возвращаем данные графика в JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse(chart_data)

    # --- Данные для остальных статистик (фильтрация по месяцу и году) ---
    deals_details = Deals.objects.all()
    if month and year:
        deals_details = deals_details.filter(date__month=int(month), date__year=int(year))
    elif month:
        deals_details = deals_details.filter(date__month=int(month))
    elif year:
        deals_details = deals_details.filter(date__year=int(year))

    suppliers_income = deals_details.values('supplier').annotate(total_income_loss=Sum('total_income_loss'))
    suppliers_income_dict = {
        contact.company.name: float(entry['total_income_loss'] or 0)
        for entry in suppliers_income
        for contact in Contact.objects.filter(company__id=entry['supplier'])
    }

    occ11_filter = Q(grade__iexact="OCC11") | Q(grade__iexact="OCC 11") | Q(grade__iexact="Loose OCC") | \
                   Q(grade__iexact="OCC 11 Bale String") | Q(grade__iexact="Printers Offcuts") | Q(
        grade__iexact="Stock Rolls")

    total_deals = deals_details.count()
    total_sale = deals_details.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_pallets = deals_details.aggregate(Sum('shipped_pallets'))['shipped_pallets__sum'] or 0
    transportation_fee = deals_details.aggregate(Sum('transport_cost'))['transport_cost__sum'] or 0
    suppliers_total = deals_details.aggregate(Sum('supplier_total'))['supplier_total__sum'] or 0
    mt_occ11 = deals_details.filter(occ11_filter).aggregate(Sum('received_quantity'))['received_quantity__sum'] or 0
    mt_plastic = deals_details.filter(grade="Flexible Plastic").aggregate(Sum('received_quantity'))['received_quantity__sum'] or 0
    mt_mixed_containers = deals_details.filter(grade="Mixed Container").aggregate(Sum('received_quantity'))['received_quantity__sum'] or 0
    income = deals_details.aggregate(Sum('total_income_loss'))['total_income_loss__sum'] or 0

    company_pallets = CompanyPallets.objects.select_related('company_name')

    # Обработка формы обновления паллет
    if request.method == 'POST' and 'update_pallets' in request.POST:
        for pallet in company_pallets:
            new_pallet_count = request.POST.get(f"pallets_{pallet.id}")
            new_cages_count = request.POST.get(f"cages_{pallet.id}")  # Получаем данные клеток
            new_bags_count = request.POST.get(f"bags_{pallet.id}")

            if new_pallet_count is not None:
                pallet.pallets_count = int(new_pallet_count)
            if new_cages_count is not None:
                pallet.cages_count = int(new_cages_count)

            if new_bags_count is not None:
                pallet.bags_count = int(new_bags_count)

            pallet.save()
        return HttpResponseRedirect(request.path)

    years = Deals.objects.annotate(year=ExtractYear('date')).values_list('year', flat=True).distinct()
    months = range(1, 13)

    # Объединяем все данные в один context
    context = {
        'chart_data': chart_data,
        'year': year,
        'suppliers_income': suppliers_income_dict,
        'total_deals': total_deals,
        'total_sale': total_sale,
        'total_pallets': total_pallets,
        'transportation_fee': transportation_fee,
        'suppliers_total': suppliers_total,
        'mt_occ11': mt_occ11,
        'mt_plastic': mt_plastic,
        'mt_mixed_containers': mt_mixed_containers,
        'income': income,
        'company_pallets': company_pallets,
        'month': month,
        'years': sorted(years),
        'months': months,
    }

    return render(request, 'crm/sales_analytics.html', context)


    # Получаем доступные года и месяцы
    years = Deals.objects.annotate(year=ExtractYear('date')).values_list('year', flat=True).distinct()
    months = range(1, 13)

    context = {
        'suppliers_income': suppliers_income_dict,
        'total_deals': total_deals,
        'total_sale': total_sale,
        'total_pallets': total_pallets,
        'transportation_fee': transportation_fee,
        'suppliers_total': suppliers_total,
        'mt_occ11': mt_occ11,
        'mt_plastic': mt_plastic,
        'mt_mixed_containers': mt_mixed_containers,
        'income': income,
        'company_pallets': company_pallets,
        'month': month,
        'year': year,
        'years': sorted(years),
        'months': months,
    }



def add_contact_material(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        # Создаем экземпляр формы с данными из POST-запроса
        form = ContactMaterialForm(request.POST, request.FILES)

        # Проверяем, что форма валидна
        if form.is_valid():
            # Если форма валидна, сохраняем материал
            contact_material = form.save(commit=False)
            contact_material.contact = contact  # Связываем материал с контактом
            contact_material.save()

            # После сохранения редиректим на страницу с контактами
            return redirect('view_contact', id = contact.id)

    else:
        # Если метод запроса GET, создаем пустую форму
        form = ContactMaterialForm()

    # Рендерим шаблон с формой и данными о контакте
    return render(request, 'crm/add_contact_material.html', {'form': form, 'contact': contact})


def edit_contact_material(request, pk):
    contact_material = get_object_or_404(ContactMaterial, pk=pk)

    if request.method == 'POST':
        form = ContactMaterialForm(request.POST, instance=contact_material)
        if form.is_valid():
            form.save()
            return redirect('view_contact', id=contact_material.contact.id)
    else:
        form = ContactMaterialForm(instance=contact_material)

    return render(request, 'crm/edit_contact_material.html', {'form': form, 'contact_material': contact_material})


#REPORTS

def report_list(request):
    return render(request, 'crm/report_list.html')


def company_report(request):
    # Получаем текущий месяц и год
    now = datetime.now()
    current_month = now.month
    current_year = now.year

    # Получаем параметры фильтра из запроса
    selected_company_id = request.GET.get('company', '')
    month = request.GET.get('month', '') or str(current_month).zfill(2)
    year = request.GET.get('year', '') or str(current_year)

    # Фильтрация сделок
    deals = Deals.objects.all()

    if selected_company_id:
        # Фильтруем по поставщикам и покупателям
        deals = deals.filter(Q(supplier__id=int(selected_company_id)))

    if month and year:
        deals = deals.filter(date__month=int(month), date__year=int(year))
    elif month:  # Если указан только месяц
        deals = deals.filter(date__month=int(month))
    elif year:  # Если указан только год
        deals = deals.filter(date__year=int(year))

    # Итоги
    total_amount_supplier = deals.aggregate(Sum('supplier_total'))['supplier_total__sum'] or 0


    # Список компаний для выбора в фильтре
    companies = Company.objects.filter(contacts__company_type='suppliers').distinct()

    # Получаем доступные года из базы данных для фильтрации
    years = Deals.objects.annotate(year=ExtractYear('date')).values_list('year', flat=True).distinct()

    # Список месяцев для фильтрации
    months = range(1, 13)  # Месяцы с 1 по 12

    # Контекст для шаблона
    context = {
        'deals': deals,
        'total_amount_supplier': total_amount_supplier,
        'companies': companies,
        'selected_company_id': int(selected_company_id) if selected_company_id.isdigit() else None,
        'month': month,
        'year': year,
        'years': sorted(years),  # Сортируем список лет для удобства
        'months': months,  # Месяцы для выпадающего списка
    }
    return render(request, 'crm/company_report.html', context)



def export_company_report_pdf(request):
    # --- импорты ---
    from io import BytesIO                                      # буфер для PDF
    from datetime import datetime                                # даты
    from django.http import HttpResponse                         # ответ
    from reportlab.lib.pagesizes import A4                       # размер страницы
    from reportlab.pdfgen import canvas                          # низкоуровневый canvas
    from reportlab.platypus import Table, TableStyle             # таблица
    from reportlab.lib import colors                             # цвета
    from reportlab.lib.utils import ImageReader                  # изображение
    import os                                                    # пути
    from django.conf import settings                             # BASE_DIR
    from django.utils.text import slugify                        # безопасное имя файла

    # --- входные параметры фильтров ---
    selected_company_id = request.GET.get('company', '')         # выбранная компания
    month = request.GET.get('month', '')                         # месяц
    year = request.GET.get('year', '')                           # год

    # --- данные ---
    now = datetime.now()                                         # текущая дата
    deals = Deals.objects.all()                                  # все сделки
    if selected_company_id:                                      # фильтр по компании
        deals = deals.filter(supplier__id=int(selected_company_id))
    if month:                                                    # фильтр по месяцу
        deals = deals.filter(date__month=int(month))
    if year:                                                     # фильтр по году
        deals = deals.filter(date__year=int(year))

    deals = deals.order_by('date', 'id')                         # стабильный порядок
    first_deal = deals.first()                                   # для шапки клиента

    # --- подготовка табличных данных ---
    data = [["Date", "Grade", "Net (MT)", "Price ($/MT)", "Amount ($)", "Scale Ticket"]]  # заголовки
    for d in deals:                                              # строки
        data.append([
            d.date.strftime("%Y-%m-%d"),
            str(d.grade),
            f"{d.received_quantity:.3f}",
            f"{d.supplier_price:.2f}",
            f"{d.supplier_total:.2f}",
            str(d.scale_ticket or "")
        ])

    # --- итоги по всему списку ---
    total_net = sum(d.received_quantity for d in deals)          # суммарный вес
    total_amount = sum(d.supplier_total for d in deals)          # суммарная сумма

    # --- подготовка PDF ---
    buffer = BytesIO()                                           # буфер
    pdf = canvas.Canvas(buffer, pagesize=A4)                     # canvas
    PAGE_W, PAGE_H = A4                                          # ширина/высота
    M_L, M_R, M_T, M_B = 30, 30, 40, 40                          # поля (мм в поинтах ≈ уже поинты)
    usable_w = PAGE_W - M_L - M_R                                # полезная ширина

    # --- ресурсы (логотип) ---
    logo_path = os.path.join(settings.BASE_DIR, 'crm', 'static', 'crm', 'images', 'log.png')
    # --- хелперы рендера ---
    def draw_header():
        """Рисует логотип/адрес/заголовок и блок с Customer. Возвращает y-координату НИЗА customer-блока."""
        y = PAGE_H - 60                                        # старт сверху
        # логотип
        if os.path.exists(logo_path):
            pdf.drawImage(ImageReader(logo_path), M_L, y - 50, width=70, height=70, mask='auto')

        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawCentredString(PAGE_W / 2, y, "Shipment Summary")

        # адрес / название справа
        pdf.setFont("Helvetica-Bold", 11)
        pdf.setFillColor(colors.darkblue)
        pdf.drawRightString(PAGE_W - M_R, y, "Local to Global Recycling Inc.")

        pdf.setFont("Helvetica", 9)
        pdf.setFillColor(colors.HexColor("#555555"))
        pdf.drawRightString(PAGE_W - M_R, y - 14, "19090 Lougheed Hwy.")
        pdf.drawRightString(PAGE_W - M_R, y - 26, "Pitt Meadows, BC V3Y 2M6")
        pdf.drawRightString(PAGE_W - M_R, y - 38, "wastepaperbrokers.com")

        # разделительная линия
        pdf.setStrokeColor(colors.HexColor("#aaaaaa"))
        pdf.setLineWidth(0.5)
        pdf.line(M_L, y - 50, PAGE_W - M_R, y - 50)
        # Собираем строки Customer
        lines = []  # список строк
        if first_deal and first_deal.supplier:
            lines.append(first_deal.supplier.name)
            contact = first_deal.supplier.contacts.filter(address__isnull=False).first()  # контакт с адресом
            if contact and contact.address:  # если адрес есть
                lines.extend(
                    [ln.strip() for ln in contact.address.strip().split('\n') if ln.strip()])  # добавляем строки адреса
            else:
                lines.append("Address not available")  # заглушка адреса
        else:
            lines.append("Unknown")  # если данных нет

        pdf.setFont("Helvetica", 10)  # обычный шрифт для Customer строк
        y_lines = PAGE_H - 135  # стартовая Y для строк адреса
        for ln in lines:  # выводим по строкам
            pdf.drawString(M_L + 55, y_lines, ln)  # печатаем строку
            y_lines -= 14  # сдвигаем вниз на 14pt

        return y_lines - 12  # немного отступа под блоком

    def table_style():
        """Единый стиль таблицы (повторяется на каждой порции)."""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ])

    # ===== Хелпер: перенос страницы при нехватке места =====
    def ensure_space_or_new_page(y_cur, need=40):
        if y_cur - need < M_B:  # если не хватает места вниз
            pdf.showPage()  # создаём новую страницу
            return draw_header()  # рисуем шапку и возвращаем новую нижнюю Y
        return y_cur  # иначе возвращаем текущую Y

    # ===== Хелпер: пагинация таблицы (рисует порциями) =====
    def draw_table_paginated(all_rows, start_y):
        """
        all_rows: список включая первую строку-заголовок.
        Рисуем постранично: на каждой странице один заголовок + часть строк данных.
        """
        col_widths = [80, 140, 60, 80, 80, 90]  # твои 6 колонок
        header = all_rows[0]  # строка-заголовок
        rows = all_rows[1:]  # только данные

        i, y_cur = 0, start_y

        while i < len(rows):
            # 1) подобрать сколько строк данных влезет вместе с заголовком
            low, high, fit = 1, len(rows) - i, 0  # минимум 1 строка данных
            while low <= high:
                mid = (low + high) // 2
                trial = [header] + rows[i:i + mid]
                t_try = Table(trial, colWidths=col_widths);
                t_try.setStyle(table_style())
                w_try, h_try = t_try.wrap(usable_w, 0)
                if y_cur - h_try >= M_B:
                    fit = mid
                    low = mid + 1
                else:
                    high = mid - 1

            # 2) если даже 1 строка не влезла — новая страница и шапка
            if fit == 0:
                pdf.showPage()
                y_cur = draw_header()
                continue

            # 3) рисуем порцию: один заголовок + fit строк данных
            chunk = [header] + rows[i:i + fit]
            t = Table(chunk, colWidths=col_widths);
            t.setStyle(table_style())
            w, h = t.wrap(usable_w, 0)
            x = M_L + (usable_w - w) / 2  # центрируем в полезной ширине
            t.drawOn(pdf, x, y_cur - h)
            y_cur -= h + 10
            i += fit

            # 4) мало места — сразу новая страница с шапкой
            if i < len(rows) and y_cur < M_B + 60:
                pdf.showPage()
                y_cur = draw_header()

        return y_cur

    # ===== Старт страницы 1: рисуем шапку и таблицу с переносами =====
    y = draw_header()  # рисуем шапку, получаем стартовую Y
    y = draw_table_paginated(data, y - 10)  # рисуем таблицу (ниже шапки)

    # ===== Итоги под таблицей (с переносом при необходимости) =====
    y -= 40
    y = ensure_space_or_new_page(y, need=40)  # проверка места под блок итогов
    pdf.setFont("Helvetica-Bold", 12)  # жирный 12pt
    pdf.drawString(M_L, y, f"Net Weight: {total_net:,.2f} MT")  # общий Net
    y -= 18  # сдвиг вниз
    pdf.drawString(M_L, y, f"Revenue: ${total_amount:.2f}")  # общий Amount
    y -= 24  # вернём, где остановились

    # агрегируем
    grade_summary = {}                                            # словарь агрегатов
    for d in deals:
        key = (str(d.grade), float(d.supplier_price))
        if key not in grade_summary:
            grade_summary[key] = {"net": 0.0, "amount": 0.0}
        grade_summary[key]["net"] += float(d.received_quantity)
        grade_summary[key]["amount"] += float(d.supplier_total)

    # сортируем для стабильности вывода
    items = sorted(grade_summary.items(), key=lambda kv: (kv[0][0], kv[0][1]))

    pdf.setFont("Helvetica", 10)
    for (grade, price), vals in items:
        line = f"{grade} (${price:.2f}) – {vals['net']:.2f} MT – ${vals['amount']:,.2f}"
        y = ensure_space_or_new_page(y, need=16)                  # перенос по мере необходимости
        pdf.drawString(M_L, y, line)
        y -= 14

    # --- имя файла ---
    raw_name = first_deal.supplier.name if first_deal and first_deal.supplier else "Unknown"
    safe_name = slugify(raw_name)
    month_str = datetime.strptime(month, "%m").strftime("%b") if month else now.strftime("%b")
    year_str = year if year else now.strftime("%Y")
    filename = f"L2G_{safe_name}_Supply_List_{month_str}_{year_str}.pdf"

    # --- финал ---
    pdf.save()                                                    # сохраняем ОДИН раз
    buffer.seek(0)                                                # в начало буфера
    response = HttpResponse(buffer, content_type='application/pdf')  # отдаём PDF
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response







def get_deal_by_ticket(request):
    ticket_number = request.GET.get('ticket_number', None)
    print(f"Получен запрос на поиск Scale Ticket: {ticket_number}")

    if not ticket_number:
        return JsonResponse({'success': False, 'error': 'No ticket number provided'})

    try:
        deal = Deals.objects.filter(scale_ticket=ticket_number).first()
        if not deal:
            print("Ошибка: Сделка не найдена")
            return JsonResponse({'success': False, 'error': 'No deal found for this ticket number'}, status=404)

        print(f"Найдена сделка: {deal}")

        return JsonResponse({
            'success': True,
            'deal': {
                'id': deal.id,
                'date': deal.date.strftime('%Y-%m-%d'),
                'received_quantity': float(deal.received_quantity),
                'received_pallets': deal.received_pallets,
                'supplier_name': deal.supplier.name if deal.supplier else "",
                'grade': deal.grade,
            }
        })
    except Exception as e:
        import traceback
        print("Ошибка сервера:", traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)







def export_scale_ticket_pdf(request):
    # ===== Импорты (если нет выше) =====
    # from io import BytesIO                                   # буфер для PDF
    # from datetime import datetime                            # даты/время
    # from decimal import Decimal                              # точная арифметика
    # import os                                                # пути/каталоги
    # from django.http import HttpResponse                     # HTTP-ответ
    # from reportlab.lib.pagesizes import A4                   # A4 как в supply list
    # from reportlab.pdfgen import canvas                      # canvas PDF
    # from reportlab.platypus import Table, TableStyle         # таблицы
    # from reportlab.lib import colors                         # цвета
    # from reportlab.lib.utils import ImageReader              # лого
    # from django.conf import settings                         # BASE_DIR, MEDIA_ROOT
    # from .models import Deals                                # модель сделок
    # from .utils import sanitize_filename                     # безопасные имена

    # ===== Хелпер: безопасное Decimal =====
    def D(val, default='0'):
        # # Приводит значение к Decimal, устойчиво к None/пустым строкам/числам
        try:
            if val is None or val == '':
                return Decimal(default)
            return Decimal(str(val))
        except Exception:
            return Decimal(default)

    # ===== Параметры запроса =====
    ticket_number = request.GET.get('ticket_number', None)                 # номер тикета из GET
    if not ticket_number:                                                  # проверка наличия номера
        return HttpResponse("⚠️ No ticket number provided.", status=400)

    # ===== Данные из БД =====
    deals = Deals.objects.filter(scale_ticket=ticket_number)               # все сделки по тикету
    if not deals.exists():                                                 # если пусто — 404
        return HttpResponse("⚠️ No deals found for this ticket number.", status=404)
    first_deal = deals.first()                                             # первая для шапки/дат

    # ===== Дата-источник =====
    date_src   = getattr(first_deal, "date", None) or datetime.today().date()  # дата сделки/сегодня
    year       = date_src.strftime("%Y")                                    # '2025'
    month_num  = date_src.strftime("%m")                                    # '09'
    month = date_src.strftime("%B")                                    # 'September'
    month_dir  = f"{year}-{month_num}"                                      # '2025-09' для папок


    # ===== Входные параметры формы =====
    licence_plate = request.GET.get('licence_plate', "N/A")                 # госномер авто
    tare_weight   = D(request.GET.get('tare_weight', '5170'))               # тара, кг (Decimal)
    deal_time     = request.GET.get('time') or datetime.now().strftime('%H:%M')  # время HH:MM

    # ===== Итоги веса (как Decimal) =====
    # # received_quantity хранится в МТ → переводим в кг (*1000)
    total_material_weight = sum(D(d.received_quantity) * D(1000) for d in deals)   # кг материалов
    total_pallets_weight  = sum(D(d.received_pallets) * D(15) for d in deals)      # кг паллет (15 кг/шт)
    net_weight   = total_material_weight + total_pallets_weight                     # нетто (кг)
    gross_weight = tare_weight + net_weight                                        # брутто (кг)

    # ===== Форматирование строк (с разделителями тысяч) =====
    gross_weight_str = f"{gross_weight:,.1f} KG"                                   # '5,195.0 KG'
    tare_weight_str  = f"{tare_weight:,.1f} KG"
    net_weight_str   = f"{net_weight:,.1f} KG"

    # ===== Подготовка PDF (как в supply list) =====
    buffer = BytesIO()                                                             # буфер в памяти
    pdf = canvas.Canvas(buffer, pagesize=A4)                                       # формат A4
    PAGE_W, PAGE_H = A4                                                            # ширина/высота
    M_L, M_R, M_T, M_B = 30, 30, 40, 40                                           # поля страницы
    usable_w = PAGE_W - M_L - M_R                                                  # полезная ширина
    logo_path = os.path.join(settings.BASE_DIR, 'crm', 'static', 'crm', 'images', 'log.png')  # лого
    INK = colors.HexColor('#141b2d')

    # ===== Цвета/хелперы для футера =====
    TEXT_MUTED = colors.HexColor("#555555")  # приглушённый текст

    def hr(y):
        """Горизонтальная тонкая линия-разделитель"""
        pdf.setStrokeColor(colors.HexColor("#aaaaaa"))
        pdf.setLineWidth(0.5)
        pdf.line(M_L, y, PAGE_W - M_R, y)

    def safe_supplier_address_lines():
        # # Собираем строки адреса поставщика из первого контакта с address, если есть
        lines = []
        try:
            supplier = getattr(first_deal, "supplier", None)  # # Компания-поставщик
            if supplier:
                lines.append(supplier.name or "Unknown")  # # Имя компании
                # Пытаемся найти контакт с адресом
                contact = getattr(supplier, "contacts", None)
                if contact:
                    c = contact.filter(address__isnull=False).first()
                    if c and c.address:
                        # Разбиваем адрес по строкам
                        addr_lines = [ln.strip() for ln in str(c.address).split('\n') if ln.strip()]
                        if addr_lines:
                            lines.extend(addr_lines)
                        else:
                            lines.append("Address not available")
                else:
                    lines.append("Address not available")
            else:
                lines.append("Unknown")
        except Exception:
            lines = ["Unknown"]
        return lines

    # ===== Хедер  =====
    def draw_header():
        y = PAGE_H - 60                                                            # старт по Y для шапки

        if os.path.exists(logo_path):                                              # лого слева
            pdf.drawImage(ImageReader(logo_path), M_L, y - 50, width=70, height=70, mask='auto')

        pdf.setFont("Helvetica-Bold", 16)                                          # заголовок по центру
        pdf.drawCentredString(PAGE_W / 2, y, "Scale Ticket")

        pdf.setFont("Helvetica-Bold", 11)                                          # реквизиты справа
        pdf.setFillColor(colors.darkblue)
        pdf.drawRightString(PAGE_W - M_R, y, "Local to Global Recycling Inc.")
        pdf.setFont("Helvetica", 9)
        pdf.setFillColor(colors.HexColor("#555555"))
        pdf.drawRightString(PAGE_W - M_R, y - 14, "19090 Lougheed Hwy.")
        pdf.drawRightString(PAGE_W - M_R, y - 26, "Pitt Meadows, BC V3Y 2M6")
        pdf.drawRightString(PAGE_W - M_R, y - 38, "wastepaperbrokers.com")

        pdf.setStrokeColor(colors.HexColor("#aaaaaa"))                             # разделительная линия
        pdf.setLineWidth(0.5)
        pdf.line(M_L, y - 50, PAGE_W - M_R, y - 50)

        # Блок «Ticket / Date / Time / Month» слева
        y_info = PAGE_H - 160
        pdf.setFillColor(colors.black)
        pdf.setFont("Helvetica-Bold", 10)
        pdf.drawString(M_L, y_info, f"Scale Ticket #: {ticket_number}")            # номер тикета
        pdf.setFont("Helvetica", 10)
        pdf.drawString(M_L, y_info - 16, f"Date:  {date_src.strftime('%Y-%m-%d')}")# дата
        pdf.drawString(M_L, y_info - 30, f"Time:  {deal_time}")                    # время


        # Блок «Customer»
        pdf.setFont("Helvetica-Bold", 10)
        pdf.setFillColor(INK)
        pdf.drawString(M_L, y_info - 68, "Customer:")
        pdf.setFont("Helvetica", 10)

        cust_lines = safe_supplier_address_lines()  # <-- имя + строки адреса
        y_lines = y_info - 84
        for ln in cust_lines:
            pdf.drawString(M_L + 12, y_lines, str(ln).strip())
            y_lines -= 14


        # Правый блок с чипами весов/табличками (простым текстом, чтобы не усложнять)
        r_x = PAGE_W - M_R - 220
        pdf.setFont("Helvetica", 10)
        pdf.drawRightString(PAGE_W - M_R, y_info,     f"Licence: {licence_plate}") # номер авто
        pdf.drawRightString(PAGE_W - M_R, y_info - 16, f"Gross:   {gross_weight_str}") # брутто
        pdf.drawRightString(PAGE_W - M_R, y_info - 30, f"Tare:    {tare_weight_str}")  # тара
        pdf.drawRightString(PAGE_W - M_R, y_info - 44, f"Net:     {net_weight_str}")   # нетто
        pdf.drawRightString(PAGE_W - M_R, y_info - 58, f"Pallets #: {int(D(getattr(first_deal,'received_pallets',0)))}") # паллеты

        return min(y - 12, y_info - 120)                                      # вернуть нижнюю границу Y

    def draw_footer():
        # Нижняя линия + подпись всегда на одинаковой высоте
        hr(M_B + 18)
        pdf.setFont("Helvetica", 8);
        pdf.setFillColor(TEXT_MUTED)
        pdf.drawString(M_L, M_B + 6, "Thank you for recycling responsibly.")
        pdf.drawRightString(PAGE_W - M_R, M_B + 6, "Generated by ADAM CRM")



    # ===== Единый стиль таблицы (как в supply list) =====
    def table_style():
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),                           # фон заголовка
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),                      # текст заголовка
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),                                  # выравнивание по центру
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),                        # жирный заголовок
            ('FONTSIZE', (0, 0), (-1, 0), 9),                                       # кегль заголовка
            ('FONTSIZE', (0, 1), (-1, -1), 9),                                      # кегль ячеек
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),                                  # отступ в заголовке
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),                          # сетка 0.5pt
        ])

    # ===== Переход на новую страницу при нехватке места =====
    def ensure_space_or_new_page(y_cur, need=40):
        if y_cur - need < M_B:
            draw_footer()
            pdf.showPage()                                                          # новая страница
            return draw_header()                                                    # рисуем шапку заново
        return y_cur                                                                # иначе вернуть текущую Y

    # ===== Пагинация таблицы =====
    def draw_table_paginated(all_rows, start_y):
        col_widths = [140, 100, 100, 110]                                          # ширины колонок
        header = all_rows[0]                                                        # шапка
        rows = all_rows[1:]                                                         # данные
        i, y_cur = 0, start_y                                                       # индекс и текущая Y
        while i < len(rows):                                                        # пока не вывели все строки
            # бинпоиск: сколько строк влезет
            low, high, fit = 1, len(rows) - i, 0
            while low <= high:
                mid = (low + high) // 2
                trial = [header] + rows[i:i + mid]
                t_try = Table(trial, colWidths=col_widths); t_try.setStyle(table_style())
                w_try, h_try = t_try.wrap(usable_w, 0)
                if y_cur - h_try >= M_B:
                    fit = mid; low = mid + 1
                else:
                    high = mid - 1
            if fit == 0:                                                            # совсем не влезает
                pdf.showPage(); y_cur = draw_header(); continue                     # новая страница и снова
            # рисуем порцию
            chunk = [header] + rows[i:i + fit]
            t = Table(chunk, colWidths=col_widths); t.setStyle(table_style())
            w, h = t.wrap(usable_w, 0)
            x = M_L + (usable_w - w) / 2                                           # центрируем в полях
            t.drawOn(pdf, x, y_cur - h)
            y_cur -= h + 10                                                         # опускаем курсор
            i += fit                                                                # смещаем индекс
            if i < len(rows) and y_cur < M_B + 60:                                  # если на странице мало места
                pdf.showPage(); y_cur = draw_header()                               # – новая страница
        return y_cur                                                                # вернуть нижнюю границу

    # ===== Подготовка данных таблицы =====
    data = [['MATERIAL', 'WEIGHT (KG)', 'PRICE ($/KG)', 'AMOUNT ($)']]             # шапка
    total_amount = Decimal('0')                                                     # итог в Decimal
    for d in deals:                                                                 # по сделкам
        received_kg  = D(d.received_quantity) * D(1000)                             # вес в кг
        sup_price_kg = D(d.supplier_price) / D(1000)                                # $/кг
        amount       = received_kg * sup_price_kg                                   # сумма
        total_amount += amount                                                      # к итогу
        data.append([
            str(d.grade),                                                           # материал
            f"{received_kg:,.1f}",                                                  # вес
            f"{sup_price_kg:.4f}",                                                  # цена/кг (4 знака)
            f"{amount:,.2f}",                                                       # сумма
        ])

    if total_pallets_weight > 0:                                                    # добавить строку паллет
        data.append(['Pallets (weight)', f"{total_pallets_weight:,.1f}", '', ''])   # вес паллет

    # ===== Рендер шапки и таблицы =====
    y = draw_header()                                                               # рисуем шапку
    y = draw_table_paginated(data, y - 10)                                          # рисуем таблицу с переносами

    # ===== Итог под таблицей =====
    y = ensure_space_or_new_page(y, need=30)                                        # место под итог
    pdf.setFont("Helvetica-Bold", 12)                                               # жирный шрифт
    pdf.drawString(M_L, y - 18, f"Revenue: ${total_amount:,.2f}")                     # ИТОГО $

    # === Футер на последней странице ===
    draw_footer()


    # ===== Пути сохранения =====
    raw_supplier_name = first_deal.supplier.name if first_deal.supplier else "Unknown Supplier"  # имя поставщика
    supplier_name = sanitize_filename(raw_supplier_name)                            # безопасное имя для FS
    directory = os.path.join(                                                       # папка: .../supplier/YYYY-MM/
        settings.MEDIA_ROOT, "reports", "scale_tickets", supplier_name, year, month
    )
    os.makedirs(directory, exist_ok=True)                                           # создаём при необходимости

    # ===== Имя файла (как указывал) =====
    filename = f"Ticket {ticket_number}-{supplier_name}-{month_dir}.pdf"            # финальное имя
    filepath = os.path.join(directory, filename)                                    # полный путь на диск

    # ===== Завершение PDF/ответ =====
    pdf.save()                                                                      # закрыть PDF
    buffer.seek(0)                                                                  # перемотка на начало
    with open(filepath, "wb") as f:                                                 # запись на диск
        f.write(buffer.getvalue())
    response = HttpResponse(buffer, content_type="application/pdf")                 # HTTP-ответ
    response["Content-Disposition"] = f'attachment; filename="{filename}"'          # совпадает с файлом
    return response

# Область доступа
SCOPES = ['https://www.googleapis.com/auth/calendar.readonly']
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CREDENTIALS_FILE = os.path.join(BASE_DIR, "c_id.json")
TOKEN_FILE = os.path.join(BASE_DIR, "token.json")


from .models import SCaleTicketStatus

def scale_ticket_browser(request):
    # 1) относительный путь
    rel = (request.GET.get("path", "") or "").strip("/")
    rel = unquote(rel)

    # 2) базовая директория
    base = (Path(settings.MEDIA_ROOT) / "reports" / "scale_tickets").resolve()
    target = (base / rel).resolve()

    if not str(target).startswith(str(base)):
        return HttpResponseBadRequest("Invalid path")
    if not target.exists():
        raise Http404("❌ Path not found")
    if not target.is_dir():
        return HttpResponseBadRequest("Not a directory")

    folders = sorted(p.name for p in target.iterdir() if p.is_dir())
    files   = sorted(p.name for p in target.iterdir() if p.is_file() and p.suffix.lower() == ".pdf")
    back_path = "/".join(rel.split("/")[:-1]) if rel else ""

    # ---------- ОПРЕДЕЛЯЕМ company ----------
    company = None
    company_id = request.GET.get("company")

    # 1) если явно передан ?company=11
    if company_id and company_id.isdigit():
        company = Company.objects.filter(id=int(company_id)).first()

    # 2) если company не нашли, пробуем по верхней папке
    if not company and rel:
        top_folder = rel.split("/", 1)[0]          # 'Inno_Foods'
        candidates = [
            top_folder,
            top_folder.replace("_", " "),          # 'Inno Foods'
        ]

        qs = Company.objects.all()
        for name in candidates:
            company = qs.filter(name__iexact=name).first()
            if company:
                break

        if not company:
            company = Company.objects.filter(unique_number__iexact=top_folder).first()

    # 3) если нашли компанию — запомним в сессии
    if company:
        request.session["current_company_id"] = company.id
    else:
        saved_id = request.session.get("current_company_id")
        if saved_id:
            company = Company.objects.filter(id=saved_id).first()

    # ---------- статусы отправок ----------
    file_statuses = {}
    for p in SCaleTicketStatus.objects.filter(sent=True).values_list("file_path", flat=True):
        k = (p or "").strip().replace("\\", "/").lstrip("/")
        if k:
            file_statuses[k] = True

    context = {
        "relative_path": rel,
        "folders": folders,
        "files": files,
        "back_path": back_path,
        "file_statuses": file_statuses,
        "company": company,
    }
    return render(request, "crm/scale_ticket_browser.html", context)

from django.core.mail import EmailMessage

@csrf_exempt
def send_scale_ticket_email(request):
    """
    Отправка Scale Ticket по email.

    Ожидаем в body (JSON):
      - path  — относительный путь внутри reports/scale_tickets
               (например: 'Local_to_Global_Recycling_Inc/2025/October/Ticket 108658-Local_to_Global_Recycling_Inc-2025-10.pdf')

    Логика:
      1) Строим абсолютный путь и проверяем, что файл существует.
      2) По первой папке (имени поставщика) находим Company.
      3) Берём всех Employee этой компании с непустым email.
      4) Отправляем email с вложением.
      5) Обновляем SCaleTicketStatus и Deals.scale_ticket_sent.
    """

    if request.method != "POST":
        return JsonResponse({"error": "Only POST allowed"}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    relative_path = (data.get("path") or "").strip()
    if not relative_path:
        return JsonResponse({"error": "Field 'path' is required"}, status=400)

    # 1) Абсолютный путь к файлу
    abs_path = os.path.join(
        settings.MEDIA_ROOT,
        "reports",
        "scale_tickets",
        relative_path
    )
    print("📎 relative_path:", relative_path)
    print("📎 abs_path:", abs_path)
    print("📎 exists?", os.path.exists(abs_path))
    if not os.path.exists(abs_path):
        return JsonResponse({"error": f"File not found: {relative_path}"}, status=404)

    # 2) Определяем компанию по верхней папке
    parts = relative_path.split("/")
    company = None
    if parts:
        supplier_folder = parts[0]  # например 'Local_to_Global_Recycling_Inc'
        candidates = [
            supplier_folder,
            supplier_folder.replace("_", " "),
        ]
        qs = Company.objects.all()
        for name in candidates:
            company = qs.filter(name__iexact=name).first()
            if company:
                break
        # запасной вариант — по unique_number
        if not company:
            company = Company.objects.filter(unique_number__iexact=supplier_folder).first()

    if not company:
        return JsonResponse({"error": "Supplier company not found for path"}, status=400)

    # 3) Берём emails сотрудников этой компании
    employees = (
        Employee.objects
        .filter(contact__company=company)
        .exclude(email__isnull=True)
        .exclude(email__exact="")
    )
    recipient_emails = [e.email for e in employees]

    if not recipient_emails:
        return JsonResponse(
            {"error": "No employee emails found for this supplier"},
            status=400,
        )

    # 4) Отправляем письмо
    try:
        email_msg = EmailMessage(
            subject="📎 Scale Ticket",
            body="Please find attached the scale ticket.",
            from_email=settings.EMAIL_HOST_USER,
            to=recipient_emails,
        )
        email_msg.attach_file(abs_path)
        email_msg.send()
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    # 5) Обновляем статус файла
    status_obj, _ = SCaleTicketStatus.objects.get_or_create(file_path=relative_path)
    status_obj.sent = True
    status_obj.sent_at = now()
    status_obj.save()

    # 6) Проставляем Deals.scale_ticket_sent = True
    try:
        filename = Path(relative_path).name   # 'Ticket 108658-Local_to_Global_Recycling_Inc-2025-10.pdf'
        ticket_number = None

        if filename.startswith("Ticket "):
            rest = filename[len("Ticket "):]  # '108658-Local_to_Global_Recycling_Inc-2025-10.pdf'
            ticket_number = rest.split("-", 1)[0].strip()

        if ticket_number:
            Deals.objects.filter(
                scale_ticket=ticket_number,
                supplier=company
            ).update(scale_ticket_sent=True)
    except Exception:
        # не критично, если парсинг номера не удался
        pass

    return JsonResponse({"success": True, "sent_to": recipient_emails})

# ID календаря
CALENDAR_ID = "dmitry@wastepaperbrokers.com"

def get_calendar_events():
    """Получает события из Google Calendar."""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    if not creds or not creds.valid:
        try:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=8081, access_type="offline", prompt="consent")
            with open(TOKEN_FILE, "w") as token:
                token.write(creds.to_json())
        except Exception as e:
            return {"error": f"Ошибка авторизации: {str(e)}"}

    try:
        service = build("calendar", "v3", credentials=creds)
        now = datetime.now(t.utc).replace(microsecond=0).isoformat() + "Z"

        events_result = service.events().list(
            calendarId=CALENDAR_ID,
            timeMin=now,
            maxResults=20,
            singleEvents=True,
            orderBy="startTime",
        ).execute()

        events = events_result.get("items", [])
    except Exception as e:
        return {"error": f"Ошибка загрузки событий: {str(e)}"}

    return events

def task_list(request):
    suppliers = Company.objects.filter(contacts__company_type="suppliers")
    buyers = Company.objects.filter(contacts__company_type="buyers")
    generate_recurring_shipments()

    context = {
        'suppliers': suppliers,
        'buyers': buyers,
        'materials_list': settings.MATERIALS_LIST,  # ✅ Передаем MATERIALS_LIST правильно
    }
    return render(request, 'crm/task_list.html', context)


def get_events(request):
    """Возвращает события в формате JSON для FullCalendar"""
    events = Event.objects.all()
    event_list = [
        {
            "id": event.id,
            "title": event.title,
            "start": event.start.strftime("%Y-%m-%dT%H:%M:%S"),
            "end": event.end.strftime("%Y-%m-%dT%H:%M:%S") if event.end else None,
            "allDay": event.all_day,
        }
        for event in events
    ]
    return JsonResponse(event_list, safe=False)

@csrf_exempt
def add_event(request):
    """Добавляет новое событие в календарь"""
    try:
        data = json.loads(request.body)
        title = data.get("title")
        start = data.get("start")

        if not title or not start:
            return JsonResponse({"error": "title и start обязательны"}, status=400)

        # ✅ Преобразуем дату
        start_datetime = datetime.fromisoformat(start)

        # ✅ Проверяем, есть ли уже таймзона, если нет — добавляем
        if start_datetime.tzinfo is None:
            start_datetime = timezone.make_aware(start_datetime)

        # ✅ Создаем событие в БД
        event = Event.objects.create(
            title=title,
            start=start_datetime,
            all_day=False
        )

        return JsonResponse({"status": "success", "event": {
            "id": event.id,
            "title": event.title,
            "start": event.start.isoformat(),
            "allDay": event.all_day
        }})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)




@csrf_exempt
def delete_event(request, event_id):
    """Удаляет событие"""
    if request.method == "DELETE":
        event = Event.objects.filter(id=event_id).first()
        if event:
            event.delete()
            return JsonResponse({"status": "deleted"})
        return JsonResponse({"error": "Event not found"}, status=404)

    return JsonResponse({"error": "Invalid request"}, status=400)



def get_grades(request):
    """
    Возвращает список всех материалов (грейдов).
    """
    return JsonResponse(list(settings.MATERIALS_LIST.keys()), safe=False)


@csrf_exempt
def add_scheduled_shipment(request):
    """
    Добавляет новую запланированную отгрузку.
    """
    if request.method == "POST":
        data = json.loads(request.body)
        try:
            supplier = Company.objects.get(id=data["supplier"])
            buyer = Company.objects.get(id=data["buyer"])

            # 🔹 Извлекаем дату и время
            date_part = data["datetime"].split("T")[0]
            time_part = data["datetime"].split("T")[1]

            # 🔹 Флаги повтора
            is_recurring = data.get("is_recurring", False)
            recurrence_type = data.get("recurrence_type", None)
            recurrence_day = data.get("recurrence_day", None)

            shipment = ScheduledShipment.objects.create(
                supplier=supplier,
                buyer=buyer,
                date=data["datetime"].split("T")[0],
                time=data["datetime"].split("T")[1],
                grade=data["grade"],
                is_recurring=data.get("is_recurring", False),
                recurrence_type=data.get("recurrence_type", None),
                recurrence_day=data.get("recurrence_day", None)
            )

            return JsonResponse({"status": "success", "shipment_id": shipment.id})
        except Company.DoesNotExist:
            return JsonResponse({"error": "Supplier or Buyer not found"}, status=400)

    return JsonResponse({"error": "Invalid request"}, status=400)


def get_scheduled_shipments(request):
    """
    Возвращает список всех запланированных отгрузок.
    """
    shipments = ScheduledShipment.objects.filter(is_done=False).order_by("date")
    shipment_list = [
        {
            "id": shipment.id,
            "date": shipment.date.strftime("%Y-%m-%d"),
            "time": shipment.time.strftime("%H:%M"),
            "supplier": shipment.supplier.name,
            "buyer": shipment.buyer.name,
            "grade": shipment.grade
        }
        for shipment in shipments
    ]
    return JsonResponse(shipment_list, safe=False)

@csrf_exempt
def delete_scheduled_shipment(request, shipment_id):
    """
    Удаляет запланированную отгрузку.
    """
    if request.method == "DELETE":
        try:
            shipment = ScheduledShipment.objects.get(id=shipment_id)
            shipment.delete()
            return JsonResponse({"status": "deleted"})
        except ScheduledShipment.DoesNotExist:
            return JsonResponse({"error": "Shipment not found"}, status=404)

    return JsonResponse({"error": "Invalid request"}, status=400)



from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


@csrf_exempt
def generate_bol_pdf(request):
    if request.method == 'POST':
        data = json.loads(request.body)

        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        spacing = 15
        y = height - 50

        # 📦 Логотип и название компании
        logo_path = os.path.join(settings.BASE_DIR, 'crm', 'static', 'crm', 'images', 'log.png')
        if os.path.exists(logo_path):
            p.drawImage(ImageReader(logo_path), 40, height - 75, width=70, height=50, mask='auto')

        # 🏢 Название компании и адрес справа
        p.setFont("Helvetica-Bold", 10)
        p.setFillColor(colors.darkblue)
        p.drawRightString(width - 40, height - 45, "Local to Global Recycling Inc.")

        p.setFont("Helvetica", 8)
        p.setFillColor(colors.black)
        p.drawRightString(width - 40, height - 58, "19090 Lougheed Hwy.")
        p.drawRightString(width - 40, height - 68, "Pitt Meadows, BC V3Y 2M6")

        # 🧾 Заголовок по центру
        p.setFont("Helvetica-Bold", 12)
        p.drawCentredString(width / 2, height - 45, "BILL OF LADING")

        # 🔹 BOL Number под заголовком (с отступом вниз)
        bol_number = data.get("bolNumber", "BOL-00000")
        p.setFont("Helvetica-Bold", 12)
        p.drawCentredString(width / 2, height - 65, f"{bol_number}")

        # 📋 Основная информация
        info_y = height - 280
        p.setFont("Helvetica", 10)

        def draw_line(label, value):
            nonlocal y
            p.drawString(40, y, f"{label}: {value}")
            y -= spacing

        freight = data.get("freightTerms", "").strip().lower()
        checkbox_line = (
            "Freight Charge Terms:\n"
            "(freight charges are prepaid unless marked otherwise)\n\n"
        
            f"Prepaid {'✓' if freight == 'prepaid' else ''}________     "
            f"Collect {'✓' if freight == 'collect' else ''}________     "
            f"3rd Party {'✓' if freight == '3rd party' else ''}________"
        )

        shipping_info = (
            f"Ship Date: {data.get('shipDate', '')}    Due Date: {data.get('dueDate', '')}\n"
            f"Carrier: {data.get('carrier', '')}\n"
            f"PO Number: {data.get('poNumber', '')}"
        )

        supp_info = (
            f"{data.get('shipFrom', '')}\n"
            f"{data.get('shipFromAddress', '')}\n"

        )

        buyer_info = (
            f"{data.get('shipTo', '')}\n"
            f"{data.get('shipToAddress', '')}\n"

        )

        references_info = (
            f"BOL # {data.get('bolNumber', '')}\n"
            f"Load # {data.get('loadNumber', '')}\n"
            f"PO Number: {data.get('poNumber', '')}"
        )

        styles = getSampleStyleSheet()

        info_data = [
            ["SHIP FROM", ""],
            [supp_info, shipping_info],
            ["SHIP TO","REFERENCES"],
            [buyer_info,references_info],
            ["THIRD PARTY FREIGHT CHARGES BILL TO: ", "FREIGHT CHARGE  TERMS: "],
            ["", checkbox_line],

        ]

        info_col_widths = [285, 285]
        info_table_width = sum(info_col_widths)
        x_info = (width - info_table_width) / 2  # Центрируем по ширине страницы

        info_table = Table(info_data, colWidths=info_col_widths)
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # SHIP FROM
            ('BACKGROUND', (0, 2), (-1, 2), colors.lightgrey),  # SHIP TO / REFERENCES
            ('BACKGROUND', (0, 4), (-1, 4), colors.lightgrey),  # THIRD PARTY / TERMS (если хочешь)

            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 2), (-1, 2), colors.black),
            ('TEXTCOLOR', (0, 4), (-1, 4), colors.black),

            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
        ]))
        info_table.wrapOn(p, width, height)
        info_table.drawOn(p, x_info, info_y)


        # 📦 Таблица товаров
        y = info_y - info_table._height + 120

        col_widths = [50, 60, 60, 50, 35, 170, 60, 40, 40]
        table_width = sum(col_widths)

        headers = ["QTY", "Handling", "PKG", "WT", "HM", "COMMODITY DESCRIPTION", "DIMS", "CLASS", "NMFC"]
        commodities = data.get('commodities', [])
        table_data = [headers]
        total_weight = 0

        for item in commodities:
            wt = float(item.get("weight", 0) or 0)
            total_weight += wt
            row = [
                item.get("qty", ""),
                item.get("handling", ""),
                item.get("pkg", ""),
                wt,
                item.get("hm", ""),
                item.get("description", ""),
                item.get("dims", ""),
                item.get("class", ""),
                item.get("nmfc", "")
            ]
            table_data.append(row)

        # 👉 добавляем строку TOTAL
        table_data.append(['', '', '', f"{total_weight:.1f}", '', 'TOTAL', '', '', ''])
        num_rows = len(table_data)

        # 👉 создаём таблицу
        commodity_table = Table(table_data, colWidths=col_widths)
        commodity_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),  # заголовки
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, 0), 0.5, colors.black),

            ('BOX', (0, 1), (-1, -1), 0.5, colors.black),  # внешняя рамка
            ('INNERGRID', (0, 1), (-1, -1), 0, colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),

            # 👉 строка TOTAL — жирная и серая
            ('BACKGROUND', (0, num_rows - 1), (-1, num_rows - 1), colors.lightgrey),
            ('FONTNAME', (0, num_rows - 1), (-1, num_rows - 1), 'Helvetica-Bold'),
            ('GRID', (0, num_rows - 1), (-1, num_rows - 1), 0.5, colors.black),
        ]))

        # 👉 позиционирование по центру страницы
        x = (width - table_width) / 2  # это ключ!
        commodity_table.wrapOn(p, width, height)
        commodity_table.drawOn(p, x, y)

    # 📦 Таблица состояния загрузки
    coll_widths = [370, 100, 100]
    table3_width = sum(coll_widths)

    # ✅ Данные чекбоксов
    trailer_loaded = data.get('trailer_loaded', '').strip().lower()
    freight_counted = data.get('freight_counted', '').strip().lower()

    # ✅ Текст с галочками
    check_trailer_load = (
        f"Trailer Loaded:\n\n"
        f"By Shipper {'✓' if trailer_loaded == 'shipper' else ''} ____\n"
        f"By Driver  {'✓' if trailer_loaded == 'driver' else ''} ____"
    )

    check_freight_counted = (
        f"Freight Counted:\n\n"
        f"By Shipper {'✓' if freight_counted == 'shipper' else ''} ____\n"
        f"By Driver  {'✓' if freight_counted == 'driver' else ''} ____"
    )

    # ✅ Заголовки и данные
    table3_data = [
        ["", "TRAILER LOADED", "FREIGHT COUNTED"],
        ["", check_trailer_load, check_freight_counted]
    ]

    # 👉 создаём таблицу
    load_table = Table(table3_data, colWidths=coll_widths)
    load_table.setStyle(TableStyle([
        ('BACKGROUND', (1, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))

    # 📌 Позиция для новой таблицы (ниже)
    y2 = y - commodity_table._height - 30  # отступ вниз
    x2 = (width - table3_width) / 2

    load_table.wrapOn(p, width, height)
    load_table.drawOn(p, x2, y2)

    # Подписи слева
    shipper1 = (
        "Shipper Signature/Date\n\n"
        "Shipper: ____________________     Date: _______________"
    )

    carrier1 = (
        "Carrier Signature/Pickup Date\n\n"
        "Carrier: ____________________     Date: _______________"
    )

    signarure2 = (

        "Receiver Signature: ________________________________\n\n"
        "Print Name: ________________________________________\n\n"
        "Exceptions: _________________________________________\n\n"
        "_____________________________________________________\n"
    )

    # 🔹 Содержимое под Pickup
    pickup_table = (
        "                Time             Shipper\n"
        "                                     Initials\n"
        "       Appt: _____           _______\n"
        "   Time In: _____           _______\n"
        "Time Out: _____           _______\n"
    )

    delivery_table = (
        "                Time             Receiver\n"
        "                                     Initials\n"
        "       Appt: _____           _______\n"
        "   Time In: _____           _______\n"
        "Time Out: _____           _______\n"
    )
    # Табличные данные
    signature_data = [
        [shipper1, carrier1],  # 1 строка: 2 графы + пустая

    ]

    # Ширины колонок — при необходимости подправь
    col_widths = [285, 285]

    # Создание таблицы
    signature_table = Table(signature_data, colWidths=col_widths)

    # Применяем объединения и стиль
    signature_table.setStyle(TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))

    y_sign = y2 - load_table._height + 20  # Или где тебе нужно
    x_sign = (width - sum(col_widths)) / 2

    signature_table.wrapOn(p, width, height)
    signature_table.drawOn(p, x_sign, y_sign)

    # Табличные данные
    time_data = [
        ["PICKUP", "DELIVERY", signarure2],
        ["Shipper Initials", "Receiver Initials", ''],
        [pickup_table, delivery_table, '']
    ]

    col1_widths = [142.5,142.5,285]

    time_table = Table(time_data, colWidths=col1_widths)

    time_table.setStyle(TableStyle([
        ('SPAN', (2, 0), (2, 2)),
        ('BACKGROUND', (0, 0), (1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),

        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),



    ]))

    y_sign = y2 - signature_table._height - 150  # Или где тебе нужно
    x_sign = (width - sum(col1_widths)) / 2

    time_table.wrapOn(p, width, height)
    time_table.drawOn(p, x_sign, y_sign)

    # ✅ Завершаем PDF
    p.showPage()
    p.save()
    buffer.seek(0)

    return HttpResponse(buffer, content_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename='BOL_{data.get('bolNumber', '00000')}.pdf'"
    })


BOL_COUNTER_FILE = os.path.join(settings.BASE_DIR, 'bol_counter.json')


@csrf_exempt
def get_bol_counters(request):
    print("📍 BOL_COUNTER_FILE path:", BOL_COUNTER_FILE)

    if request.method == 'GET':
        if not os.path.exists(BOL_COUNTER_FILE):
            with open(BOL_COUNTER_FILE, 'w') as f:
                json.dump({"bol": 1000, "load": 2000}, f)


        with open(BOL_COUNTER_FILE, 'r') as f:
            data = json.load(f)
        return JsonResponse(data)

    return JsonResponse({"error": "Invalid request"}, status=400)


@csrf_exempt
def increment_bol_counters(request):
    if request.method == 'POST':
        # ✅ Если файл не существует, создаём его с начальными значениями
        if not os.path.exists(BOL_COUNTER_FILE):
            with open(BOL_COUNTER_FILE, 'w') as f:
                json.dump({"bol": 1000, "load": 2000}, f)

        # ✅ Читаем и обновляем
        with open(BOL_COUNTER_FILE, 'r+') as f:
            data = json.load(f)
            data["bol"] += 1
            data["load"] += 1
            f.seek(0)
            json.dump(data, f, indent=2)
            f.truncate()

        return JsonResponse({"status": "updated", "bol": data["bol"], "load": data["load"]})
    else:
        return JsonResponse({"error": "Invalid request"}, status=400)




# Показать задачи, привязанные к контакту
def contact_tasks(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)
    tasks = contact.tasks.all()
    return render(request, 'crm/task_list.html', {'contact': contact, 'tasks': tasks})

# Добавить задачу
def add_task(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)

    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.contact = contact  # или task.company = contact.company если менял логику
            task.save()
            return redirect('contact_tasks', contact_id=contact.id)
    else:
        form = TaskForm()

    return render(request, 'crm/add_truck.html', {'form': form, 'contact': contact})



from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def update_stage(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            pipeline_id = data.get('id')
            new_stage = data.get('stage')
            new_order = data.get('order', 0)

            pipeline = PipeLine.objects.get(id=pipeline_id)
            pipeline.stage = new_stage
            pipeline.order = new_order
            pipeline.save()

            return JsonResponse({"status": "success", "stage": new_stage, "order": new_order})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    return JsonResponse({"error": "Invalid request"}, status=400)


from django.template.defaulttags import register
@register.filter
def get_item(dictionary, key):
    return dictionary.get(key, [])

def kanban_board(request):
    stages = ["new", "send_email", "meeting", "account", "deal"]
    pipeline = PipeLine.objects.select_related("contact", "contact__company").all()

    grouped = {stage: [] for stage in stages}
    for item in pipeline:
        grouped[item.stage].append(item)

    return render(request, "crm/kanban_board.html", {
        "stages": stages,
        "grouped_pipeline": grouped
    })


@csrf_exempt
def get_clients_grouped(request):
    if request.method == "GET":
        suppliers = Client.objects.filter(client_type='suppliers').values("id", "name")
        buyers = Client.objects.filter(client_type='buyers').values("id", "name")
        return JsonResponse({
            "suppliers": list(suppliers),
            "buyers": list(buyers)
        })

def get_companies_by_type(request):
    suppliers = Company.objects.filter(contacts__company_type="suppliers").distinct()
    buyers = Company.objects.filter(contacts__company_type="buyers").distinct()
    hauler = Company.objects.filter(contacts__company_type="hauler").distinct()

    data = {
        "suppliers": [{"id": c.id, "name": c.name} for c in suppliers],
        "buyers": [{"id": c.id, "name": c.name} for c in buyers],
        "hauler": [{"id": c.id, "name": c.name} for c in hauler],
    }
    return JsonResponse(data)


@csrf_exempt
def mark_shipment_done(request, shipment_id):
    if request.method == "POST":
        try:
            shipment = ScheduledShipment.objects.get(id=shipment_id)
            shipment.is_done = True
            shipment.save()
            return JsonResponse({"status": "done"})
        except ScheduledShipment.DoesNotExist:
            return JsonResponse({"status": "not found"}, status=404)
    return JsonResponse({"status": "invalid method"}, status=405)



def supply_list(request):
    # Получаем текущий месяц и год
    now = datetime.now()
    current_month = now.month
    current_year = now.year

    # Получаем параметры фильтра из запроса
    selected_company_id = request.GET.get('company', '')
    month = request.GET.get('month', '') or str(current_month).zfill(2)
    year = request.GET.get('year', '') or str(current_year)


    # Фильтрация сделок
    deals = Deals.objects.all()

    if selected_company_id:
        deals = deals.filter(Q(buyer__id=int(selected_company_id)))

    if month and year:
        deals = deals.filter(date__month=int(month), date__year=int(year))
    elif month:  # Если указан только месяц
        deals = deals.filter(date__month=int(month))
    elif year:  # Если указан только год
        deals = deals.filter(date__year=int(year))

    # Итоги
    total_amount_buyer = deals.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_buyer_paid = deals.aggregate(Sum('total_income_loss'))[
                           'total_income_loss__sum'] or 0  # Итог для покупателя (например, прибыль или убыток)

    # Список компаний для выбора в фильтре
    companies = Company.objects.filter(contacts__company_type='buyers').distinct()



    # Получаем доступные года из базы данных для фильтрации
    years = Deals.objects.annotate(year=ExtractYear('date')).values_list('year', flat=True).distinct()

    # Список месяцев для фильтрации
    months = range(1, 13)  # Месяцы с 1 по 12

    # Контекст для шаблона
    context = {
        'deals': deals,
        'total_amount_buyer': total_amount_buyer,
        'total_buyer_paid': total_buyer_paid,  # Передаем total_buyer_paid в шаблон
        'companies': companies,
        'selected_company_id': int(selected_company_id) if selected_company_id.isdigit() else None,
        'month': month,
        'year': year,
        'years': sorted(years),  # Сортируем список лет для удобства
        'months': months,  # Месяцы для выпадающего списка
    }
    return render(request, 'crm/supply_list.html', context)



def sanitize_filename(name):
    name = name.strip()
    name = name.replace(' ', '_')
    return re.sub(r'[<>:"/\\|?*]', '_', name)

def export_supply_list_pdf(request):
    # ===== Импорты (локально внутри вьюхи, чтобы не ломать глобальный импорт) =====
    from io import BytesIO                                   # буфер в памяти под PDF
    from datetime import datetime                            # даты/время
    from django.http import HttpResponse                     # HTTP-ответ с файлом
    from reportlab.lib.pagesizes import A4                   # размер страницы A4
    from reportlab.pdfgen import canvas                      # низкоуровневый PDF canvas
    from reportlab.platypus import Table, TableStyle         # таблицы Platypus (но без Story)
    from reportlab.lib import colors                          # цвета
    from reportlab.lib.utils import ImageReader              # загрузка картинок
    import os                                                # работа с путями
    from django.conf import settings                         # BASE_DIR и проч.
    from django.utils.text import slugify                    # безопасное имя файла

    # ===== Входные параметры фильтров =====
    selected_company_id = request.GET.get('company', '')     # id компании-покупателя
    month = request.GET.get('month', '')                     # месяц (числом, "08")
    year = request.GET.get('year', '')                       # год ("2025")
    now = datetime.now()                                     # текущая дата

    # ===== Данные из базы =====
    deals = Deals.objects.all()                              # все сделки
    if selected_company_id:                                  # фильтр по покупателю
        deals = deals.filter(buyer__id=int(selected_company_id))
    if month:                                                # фильтр по месяцу
        deals = deals.filter(date__month=int(month))
    if year:                                                 # фильтр по году
        deals = deals.filter(date__year=int(year))
    deals = deals.order_by('date', 'id')                     # стабильный порядок строк
    first_deal = deals.first()                               # первая сделка (для блока Customer)

    # ===== Подготовка табличных данных =====
    data = [["Date", "Grade", "Net (MT)", "Price ($/MT)", "Amount ($)"]]  # заголовок колонок
    for d in deals:                                           # формируем строки таблицы
        data.append([
            d.date.strftime("%Y-%m-%d"),                      # дата
            str(d.grade),                                     # грейд (строкой)
            f"{float(d.shipped_quantity):.3f}",               # net вес (MT)
            f"{float(d.buyer_price):.2f}",                    # цена за MT
            f"{float(d.total_amount):.2f}",                   # сумма
        ])

    # ===== Глобальные итоги =====
    total_net = sum(float(d.shipped_quantity) for d in deals) # общий вес
    total_amount = sum(float(d.total_amount) for d in deals)  # общий оборот $

    # ===== Буфер и canvas =====
    buffer = BytesIO()                                        # создаём буфер в памяти
    pdf = canvas.Canvas(buffer, pagesize=A4)                  # инициализируем canvas
    PAGE_W, PAGE_H = A4                                       # ширина/высота страницы

    # ===== Поля страницы и полезная ширина =====
    M_L, M_R, M_T, M_B = 30, 30, 40, 40                       # левое/правое/верх/низ, в поинтах
    usable_w = PAGE_W - M_L - M_R
     # доступная ширина контента

    # ===== Путь к логотипу =====
    logo_path = os.path.join(settings.BASE_DIR, 'crm', 'static', 'crm', 'images', 'log.png')

    # ===== Хелпер: рисует шапку + Customer, возвращает нижнюю Y =====
    def draw_header():
        y = PAGE_H - 60

        if os.path.exists(logo_path):
            pdf.drawImage(ImageReader(logo_path), M_L, y - 50, width=70, height=70, mask='auto')

        # заголовок
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawCentredString(PAGE_W / 2, y, "Shipment Summary")

        # реквизиты справа
        pdf.setFont("Helvetica-Bold", 11)
        pdf.setFillColor(colors.darkblue)
        pdf.drawRightString(PAGE_W - M_R, y, "Local to Global Recycling Inc.")

        pdf.setFont("Helvetica", 9)
        pdf.setFillColor(colors.HexColor("#555555"))
        pdf.drawRightString(PAGE_W - M_R, y - 14, "19090 Lougheed Hwy.")
        pdf.drawRightString(PAGE_W - M_R, y - 26, "Pitt Meadows, BC V3Y 2M6")
        pdf.drawRightString(PAGE_W - M_R, y - 38, "wastepaperbrokers.com")

        # разделительная линия
        pdf.setStrokeColor(colors.HexColor("#aaaaaa"))
        pdf.setLineWidth(0.5)
        pdf.line(M_L, y - 50, PAGE_W - M_R, y - 50)
        # Собираем строки Customer
        lines = []                                            # список строк
        if first_deal and first_deal.buyer:                   # если есть покупатель
            lines.append(first_deal.buyer.name)               # имя компании
            contact = first_deal.buyer.contacts.filter(address__isnull=False).first()  # контакт с адресом
            if contact and contact.address:                   # если адрес есть
                lines.extend([ln.strip() for ln in contact.address.strip().split('\n') if ln.strip()])  # добавляем строки адреса
            else:
                lines.append("Address not available")         # заглушка адреса
        else:
            lines.append("Unknown")                           # если данных нет

        pdf.setFont("Helvetica", 10)                          # обычный шрифт для Customer строк
        y_lines = PAGE_H - 135                                # стартовая Y для строк адреса
        for ln in lines:                                      # выводим по строкам
            pdf.drawString(M_L + 55, y_lines, ln)             # печатаем строку
            y_lines -= 14                                     # сдвигаем вниз на 14pt

        return y_lines - 12                                   # небольшой отступ и вернуть нижнюю границу

    # ===== Хелпер: единый стиль таблицы =====
    def table_style():
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),     # серый фон заголовка
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),# белый текст заголовка
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),            # выравнивание по центру
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # жирный шрифт заголовка
            ('FONTSIZE', (0, 0), (-1, 0), 9),                 # размер заголовка
            ('FONTSIZE', (0, 1), (-1, -1), 9),                # размер ячеек
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),            # внутренний отступ снизу в заголовке
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),    # сетка 0.5pt
        ])

    # ===== Хелпер: перенос страницы при нехватке места =====
    def ensure_space_or_new_page(y_cur, need=40):
        if y_cur - need < M_B:                                 # если не хватает места вниз
            pdf.showPage()                                     # создаём новую страницу
            return draw_header()                               # рисуем шапку и возвращаем новую нижнюю Y
        return y_cur                                           # иначе возвращаем текущую Y

    # ===== Хелпер: пагинация таблицы (рисует порциями) =====
    def draw_table_paginated(all_rows, start_y):
        col_widths = [80, 160, 80, 90, 100]  # подгони при желании
        header = all_rows[0]
        rows = all_rows[1:]

        i, y_cur = 0, start_y

        while i < len(rows):
            # подобрать, сколько строк данных влезет вместе с header
            low, high, fit = 1, len(rows) - i, 0
            while low <= high:
                mid = (low + high) // 2
                trial = [header] + rows[i:i + mid]
                t_try = Table(trial, colWidths=col_widths);
                t_try.setStyle(table_style())
                w_try, h_try = t_try.wrap(usable_w, 0)
                if y_cur - h_try >= M_B:
                    fit = mid
                    low = mid + 1
                else:
                    high = mid - 1

            if fit == 0:
                pdf.showPage()
                y_cur = draw_header()
                continue

            # рисуем порцию
            chunk = [header] + rows[i:i + fit]
            t = Table(chunk, colWidths=col_widths);
            t.setStyle(table_style())
            w, h = t.wrap(usable_w, 0)
            x = M_L + (usable_w - w) / 2  # центр в полях
            t.drawOn(pdf, x, y_cur - h)
            y_cur -= h + 10
            i += fit

            if i < len(rows) and y_cur < M_B + 60:
                pdf.showPage()
                y_cur = draw_header()

        return y_cur                                           # возвращаем нижнюю Y после таблицы

    # ===== Старт страницы 1: рисуем шапку и таблицу с переносами =====
    y = draw_header()                                          # рисуем шапку, получаем стартовую Y
    y = draw_table_paginated(data, y - 10)                     # рисуем таблицу (ниже шапки)

    # ===== Итоги под таблицей (с переносом при необходимости) =====
    y -= 40
    y = ensure_space_or_new_page(y, need=40)                   # проверка места под блок итогов
    pdf.setFont("Helvetica-Bold", 12)                          # жирный 12pt
    pdf.drawString(M_L, y, f"Net Weight: {total_net:.2f} MT")   # общий Net
    y -= 18                                                    # сдвиг вниз
    pdf.drawString(M_L, y, f"Amount Due: ${total_amount:.2f}")  # общий Amount
    y -= 24                                                    # отступ перед сводкой по грейдам

    # ===== Итоги по каждому грейду (агрегация + перенос строк при необходимости) =====
    grade_summary = {}                                         # словарь агрегатов по (grade, price)
    for d in deals:                                            # проходим сделки
        key = (str(d.grade), float(d.buyer_price))             # ключ: (грейд, цена)
        if key not in grade_summary:                           # если ключ новый
            grade_summary[key] = {"net": 0.0, "amount": 0.0}   # инициализируем
        grade_summary[key]["net"] += float(d.shipped_quantity) # суммируем вес
        grade_summary[key]["amount"] += float(d.total_amount)  # суммируем сумму $

    items = sorted(grade_summary.items(), key=lambda kv: (kv[0][0], kv[0][1]))  # сортировка по названию/цене
    pdf.setFont("Helvetica", 10)                                # обычный 10pt
    for (grade, price), vals in items:                          # пробегаем агрегаты
        line = f"{grade} (${price:.2f}) – {vals['net']:.2f} MT – ${vals['amount']:.2f}"  # формируем строку
        y = ensure_space_or_new_page(y, need=16)                # проверяем место под строку
        pdf.drawString(M_L, y, line)                            # печатаем строку
        y -= 14                                                 # шаг вниз под следующую

    # ===== Имя файла =====
    raw_name = first_deal.buyer.name if first_deal and first_deal.buyer else "Unknown"   # имя компании
    safe_name = slugify(raw_name)                                 # безопасное имя
    month_str = datetime.strptime(month, "%m").strftime("%b") if month else now.strftime("%b")  # "Aug"
    year_str = year if year else now.strftime("%Y")               # "2025"
    filename = f"L2G_{safe_name}_Supply_List_{month_str}_{year_str}.pdf"  # финальное имя файла

    # ===== Завершение PDF и ответ =====
    pdf.save()                                                    # сохраняем PDF (ОДИН раз!)
    buffer.seek(0)                                                # ставим курсор в начало буфера
    response = HttpResponse(buffer, content_type='application/pdf')  # готовим ответ
    response['Content-Disposition'] = f'attachment; filename="{filename}"'  # заголовок скачивания
    return response


from datetime import timedelta, date
from crm.models import ScheduledShipment


def generate_recurring_shipments():
    print("🚀 Функция генерации повторов ЗАПУЩЕНА")

    today = date.today()
    end_date = today + timedelta(weeks=52)  # 🔁 Генерируем на год вперёд

    # Получаем шаблоны
    recurring_shipments = ScheduledShipment.objects.filter(is_recurring=True, is_done=False)
    print(f"📦 Проверка повторяющихся отгрузок на {today} ({len(recurring_shipments)} шаблонов)")

    for shipment in recurring_shipments:
        current_date = shipment.date

        while True:
            # 🗓 Вычисляем следующую дату
            if shipment.recurrence_type == 'weekly':
                current_date += timedelta(weeks=1)
            elif shipment.recurrence_type == 'biweekly':
                current_date += timedelta(weeks=2)
            elif shipment.recurrence_type == 'monthly':
                try:
                    # Пробуем сдвинуть месяц
                    next_month = current_date.month + 1
                    year = current_date.year + (next_month - 1) // 12
                    month = (next_month - 1) % 12 + 1
                    current_date = current_date.replace(year=year, month=month)
                except ValueError:
                    print(f"⚠️ Пропущено: невозможно установить дату для {shipment}")
                    break

            if current_date > end_date:
                break

            # Проверяем, не существует ли уже такая отгрузка
            exists = ScheduledShipment.objects.filter(
                supplier=shipment.supplier,
                buyer=shipment.buyer,
                date=current_date,
                time=shipment.time,
                grade=shipment.grade
            ).exists()

            if exists:
                print(f"⏩ Уже есть: {current_date} — пропущено.")
                continue

            # Создаём новую отгрузку
            ScheduledShipment.objects.create(
                supplier=shipment.supplier,
                buyer=shipment.buyer,
                date=current_date,
                time=shipment.time,
                grade=shipment.grade,
                is_recurring=False  # 📌 Повтор создан — это не шаблон
            )

            print(f"✅ Создана отгрузка на {current_date} для {shipment.supplier} → {shipment.buyer} ({shipment.grade})")


def ai_dashboard(request):
    # Выбираем убыточные сделки
    problem_deals = Deals.objects.filter(total_income_loss__lt=0)
    deal_issues_count = problem_deals.count()

    return render(request, 'crm/ai_dashboard/ai_dashboard.html', {
        'deal_issues_count': deal_issues_count
    })

# views.py
from crm.ai_dashboard.deal_recommendations import analyze_deals

def deal_recommendations(request):
    recommendations = analyze_deals()
    return render(request, 'crm/ai_dashboard/deal_recommendations.html', {
        'recommendations': recommendations
    })

from crm.ai_dashboard.shipment_predictor import predict_shipments

def shipment_predictions(request):
    predictions = predict_shipments()
    return render(request, 'crm/ai_dashboard/shipment_predictor.html', {
        'predictions': predictions
    })

from crm.ai_dashboard.client_monitor import find_inactive_clients,INACTIVITY_DAYS_THRESHOLD

def client_monitor_view(request):
    inactive_clients = find_inactive_clients()
    return render(request, 'crm/ai_dashboard/client_monitor.html', {
        'inactive_clients': inactive_clients,
        'threshold': INACTIVITY_DAYS_THRESHOLD
    })

from crm.ai_dashboard.email_generator import generate_reminder_email

def generate_email_view(request, company_id):
    from crm.models import Company
    from crm.ai_dashboard.email_generator import generate_reminder_email

    company = Company.objects.get(id=company_id)
    email_text = generate_reminder_email(company.name)

    return render(request, 'crm/ai_dashboard/generated_email.html', {
        'company': company,
        'email_text': email_text
    })

from crm.ai_dashboard.insight_engine import get_pie_chart_data
from django.http import JsonResponse

def ai_pie_stats(request):
    data = get_pie_chart_data()
    return JsonResponse(data)

from crm.models import Deals
from django.db.models import Sum

def get_problem_suppliers():
    suppliers = (
        Deals.objects
        .filter(total_income_loss__lt=0, supplier__isnull=False)
        .values('supplier__name')
        .annotate(total_loss=Sum('total_income_loss'))
        .order_by('total_loss')
    )

    max_loss = abs(min(s['total_loss'] for s in suppliers)) or 1  # защита от деления на 0

    result = []
    for s in suppliers:
        loss = abs(s['total_loss'])
        percent = int((loss / max_loss) * 100)
        result.append({
            'name': s['supplier__name'],
            'loss': round(loss, 2),
            'percent': percent
        })

    return result


from django.utils.timezone import now

def get_supplier_monthly_stats(request):
    current_year = now().year

    # Берём сделки по году, сгруппированные по supplier и месяцу
    data = (
        Deals.objects.filter(
            date__year=current_year,
            supplier__isnull=False
        )
        .values('supplier__name', 'date__month')
        .annotate(total=Sum('total_income_loss'))
        .order_by('supplier__name', 'date__month')
    )

    # Строим структуру: {supplier_name: [month1, ..., month12]}
    result = {}
    for entry in data:
        name = entry['supplier__name']
        month = entry['date__month']
        total = round(entry['total'], 2)

        if name not in result:
            result[name] = [0] * 12

        result[name][month - 1] = total

    return JsonResponse(result)

def get_buyer_supplier_map(request):
    from crm.models import Deals
    result = {}
    qs = Deals.objects.filter(buyer__isnull=False, supplier__isnull=False).values("buyer__name", "supplier__name").distinct()

    for row in qs:
        buyer = row["buyer__name"]
        supplier = row["supplier__name"]
        result.setdefault(buyer, []).append(supplier)

    return JsonResponse(result)



from crm.ai_dashboard.insight_engine import (
    get_top_clients,
    get_worst_deals,
    get_top_suppliers,
    get_problem_suppliers,
    get_clients_with_drop,
    get_supplier_monthly_profit_and_tonnage,
    get_pie_chart_data,
    compute_kpi,
    monthly_trends_data,
)
from django.views.decorators.http import require_GET
from collections import defaultdict

def insight_dashboard(request):
    # # основные блоки для шаблона
    top_clients = get_top_clients()                         # топ клиентов по прибыли
    worst_deals = get_worst_deals()                         # худшие сделки этого месяца
    top_suppliers = get_top_suppliers()                     # топ поставщиков по прибыли
    problem_suppliers = get_problem_suppliers()             # проблемные поставщики (месяц)
    dropped_clients = get_clients_with_drop()               # клиенты с падением оборота
    kpi = compute_kpi()



    # # рендерим HTML, остальное фронт подтянет через fetch()
    return render(request, 'crm/ai_dashboard/insights.html', {
        'top_clients': top_clients,
        'worst_deals': worst_deals,
        'top_suppliers': top_suppliers,
        'problem_suppliers': problem_suppliers,
        'dropped_clients': dropped_clients,
        'kpi': kpi,

    })


@require_GET
def supplier_monthly_api(request):
    # # отдаём объединённую структуру: { "Supplier A": {"profit":[..12..], "tonnage":[..12..]}, ... }
    data = get_supplier_monthly_profit_and_tonnage()        # считаем в engine
    return JsonResponse(data, safe=True)                    # safe=True т.к. dict


@require_GET
def buyer_suppliers_api(request):
    # # строим мэппинг: покупатель -> список его поставщиков (по факту сделок)
    mapping = defaultdict(set)                              # set чтобы убрать дубликаты

    qs = (Deals.objects
          .filter(buyer__isnull=False, supplier__isnull=False)   # обе стороны заданы
          .values('buyer__name', 'supplier__name')               # берём имена
          )

    for row in qs:
        buyer = row['buyer__name'] or 'Unknown'             # подстраховка от NULL
        supplier = row['supplier__name'] or 'Unknown'       # подстраховка от NULL
        mapping[buyer].add(supplier)                        # добавляем связь

    # # превращаем set -> list для JSON
    data = {buyer: sorted(list(suppliers)) for buyer, suppliers in mapping.items()}
    return JsonResponse(data, safe=True)


@require_GET
def pie_stats_api(request):
    # # отдаём данные для пирогов: {"suppliers": {...}, "buyers": {...}}
    stats = get_pie_chart_data()                            # считает engine
    return JsonResponse(stats, safe=True)

@require_GET
def monthly_trends_api(request):
    return JsonResponse(monthly_trends_data(), safe=True)



from django.views.decorators.http import require_http_methods
import math
@require_http_methods(["GET", "POST"])
def add_truck(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)
    company = contact.company

    if contact.company_type != "hauler":
        return HttpResponse("❌ Only hauler contacts can have trucks", status=400)

    if request.method == "POST":
        max_tons = float(request.POST.get("max_tons", 0))      # максимум тонн
        base_cost = float(request.POST.get("base_cost", 0))    # базовая цена
        max_spots = int(request.POST.get('max_spots', 0))

        # 🔁 Изменить расчёт max_bales:
        if max_tons > 0 and max_spots > 0:
            max_bales = max_spots  # 💡 теперь не рассчитываем, а вводим напрямую
        else:
            max_bales = 0

        # ✅ Сохраняем
        TruckProfile.objects.create(
            company=company,
            max_bales=max_bales,
            max_tons=max_tons,
            max_spots=max_spots,
            base_cost=base_cost
        )

        return redirect("view_contact", id=contact_id)

    return render(request, "crm/add_truck.html", {"contact": contact})


def delete_truck(request, id):
    truck = get_object_or_404(TruckProfile, id=id)

    # 🧩 Получаем первый контакт этой компании (предполагаем, что он есть)
    contact = Contact.objects.filter(company=truck.company).first()
    if not contact:
        return redirect('contacts_list')  # fallback, если что-то пошло не так

    truck.delete()
    return redirect('view_contact', id=contact.id)




from django.apps import apps

def _get_staff_model():
    # Пробуем несколько типичных имён модели сотрудников
    for name in ("Employee", "Staff", "StaffContact", "CompanyEmployee"):
        try:
            return apps.get_model("crm", name)
        except LookupError:
            continue
    raise LookupError("Не нашёл модель сотрудников в приложении crm (ожидал Employee/Staff/…).")

@require_http_methods(["GET"])
def api_company_contacts(request, company_id):
    company = get_object_or_404(Company, pk=company_id)

    Staff = _get_staff_model()  # модель сотрудников (имеет FK на Company)
    # Подбираем поля имени (на случай full_name/name)
    name_field = "name"
    if not hasattr(Staff, name_field) and hasattr(Staff, "full_name"):
        name_field = "full_name"

    # Фильтруем сотрудников компании, у кого есть e-mail
    qs = (
        Staff.objects
        .filter(company_id=company.id)
        .filter(~Q(email=None))
        .exclude(email__exact="")
    )

    data = [
        {
            "id": s.id,
            "name": getattr(s, name_field, ""),
            "email": s.email,
        }
        for s in qs
    ]
    return JsonResponse({"contacts": data})


@require_http_methods(["POST"])
def api_send_report_email(request):
    """
    Отправка scale ticket’а по выбранным сотрудникам.
    Ожидаем:
      - company_id      (пока не используем, но оставляем на будущее)
      - subject
      - body
      - attachment_relative  (относительный путь внутри reports/scale_tickets)
      - contact_ids          (JSON-список id сотрудников Employee)
    """
    company_id = request.POST.get("company_id")  # на будущее, не обязательно
    subject = request.POST.get("subject", "Scale Ticket")
    body = request.POST.get("body", "")
    rel_path = request.POST.get("attachment_relative")
    ids_raw = request.POST.get("contact_ids")

    if not rel_path:
        return HttpResponseBadRequest("attachment_relative is required")

    # --- парсим список id сотрудников (contact_ids) ---
    import json
    try:
        contact_ids = json.loads(ids_raw) if ids_raw else []
    except Exception:
        contact_ids = []

    if not contact_ids:
        return HttpResponseBadRequest("No recipients selected")

    # --- достаём сотрудников и их email ---
    employees = (
        Employee.objects
        .filter(id__in=contact_ids)
        .exclude(email__isnull=True)
        .exclude(email__exact="")
    )
    to_emails = [e.email for e in employees]

    if not to_emails:
        return HttpResponseBadRequest("No valid emails for selected employees")

    # --- строим абсолютный путь к файлу ---
    abs_path = os.path.join(
        settings.MEDIA_ROOT,
        "reports",
        "scale_tickets",
        rel_path
    )
    if not os.path.exists(abs_path):
        return HttpResponseBadRequest(f"File not found: {rel_path}")

    # --- шлём письмо ---
    msg = EmailMessage(
        subject=subject,
        body=body,
        from_email=settings.EMAIL_HOST_USER,
        to=to_emails,
    )
    msg.attach_file(abs_path)
    msg.send(fail_silently=False)

    # --- отмечаем статус отправки в SCaleTicketStatus ---
    status_obj, _ = SCaleTicketStatus.objects.get_or_create(file_path=rel_path)
    status_obj.sent = True
    status_obj.sent_at = timezone.now()
    status_obj.save()

    return JsonResponse({"ok": True, "sent_to": to_emails})




from django.utils.dateparse import parse_date
from django.utils.dateparse import parse_datetime

# -----------------------------
# Helpers
# -----------------------------

def _as_int(v):
    try:
        return int(v)
    except Exception:
        return None


def _clean_str(v):
    return (v or "").strip()


def _parse_date_or_none(v):
    v = _clean_str(v)
    if not v:
        return None
    return parse_date(v)


def _parse_dt_or_none(v):
    v = _clean_str(v)
    if not v:
        return None
    # Accept ISO datetime if you ever pass it
    dt = parse_datetime(v)
    return dt


def _lane_from_payload(v):
    """
    Accept lane as:
    - lane_id (int or str int) -> fetch
    - lane name (string) -> get_or_create
    """
    if v in ("", None):
        return None

    lane_id = _as_int(v)
    if lane_id:
        return ExportLane.objects.filter(id=lane_id).first()

    name = _clean_str(v)
    if not name:
        return None

    lane_obj, _ = ExportLane.objects.get_or_create(name=name)
    return lane_obj


def _schedule_from_payload(v):
    """
    Accept schedule as:
    - schedule_id (int or str int) -> fetch
    """
    if v in ("", None):
        return None

    sch_id = _as_int(v)
    if not sch_id:
        return None

    return VesselSchedule.objects.filter(id=sch_id).first()


def _deal_from_payload(v):
    """
    If you later add deal FK to ExportShipment, keep this.
    If ExportShipment has no deal field -> DON'T call this.
    """
    if v in ("", None):
        return None
    deal_id = _as_int(v)
    if not deal_id:
        return None
    return Deals.objects.filter(id=deal_id).first()


# -----------------------------
# LIST PAGE
# -----------------------------

def export_shipments_list(request):
    exports_qs = (
        ExportShipment.objects
        .select_related("lane", "schedule", "schedule__lane", "created_by")
        .prefetch_related("documents")
        .order_by("-created_at")
    )

    lane_id     = _clean_str(request.GET.get("lane"))
    status      = _clean_str(request.GET.get("status"))
    mode        = _clean_str(request.GET.get("mode"))
    schedule_id = _clean_str(request.GET.get("schedule"))
    deal_id     = _clean_str(request.GET.get("deal"))  # ← вернул

    if lane_id:
        exports_qs = exports_qs.filter(lane_id=lane_id)
    if status:
        exports_qs = exports_qs.filter(status=status)
    if mode:
        exports_qs = exports_qs.filter(mode=mode)
    if schedule_id:
        exports_qs = exports_qs.filter(schedule_id=schedule_id)
    if hasattr(ExportShipment, "deal") and deal_id:
        exports_qs = exports_qs.filter(deal_id=deal_id)

    lanes = ExportLane.objects.filter(is_active=True).order_by("name")

    schedules = (
        VesselSchedule.objects
        .select_related("lane")
        .filter(is_active=True)
        .order_by("lane__name", "doc_cutoff_at", "erd_at", "cargo_cutoff_at")
    )

    deals = (
        Deals.objects
        .select_related("supplier", "transport_company")
        .order_by("-date")[:500]
    )

    context = {
        "shipments":      exports_qs,
        "lanes":          lanes,
        "schedules":      schedules,
        "deals":          deals,

        "mode_choices":   ExportShipment.MODE_CHOICES,
        "status_choices": ExportShipment.STATUS_CHOICES,
        "doc_type_choices": ExportDocument.DOC_TYPE_CHOICES,

        "selected_lane_id":     _as_int(lane_id) or "",
        "selected_status":      status,
        "selected_mode":        mode,
        "selected_schedule_id": _as_int(schedule_id) or "",
        "selected_deal_id":     _as_int(deal_id) or "",  # ← вернул
    }

    return render(request, "crm/export_shipments.html", context)


# -----------------------------
# DETAIL JSON (for View sidebar)
# -----------------------------

def export_shipment_detail_json(request, pk):
    export = get_object_or_404(
        ExportShipment.objects.select_related("lane", "schedule").prefetch_related("documents"),
        pk=pk,
    )

    # Deal is OPTIONAL. Only fill supplier/grade/carrier if your model has `deal` FK.
    supplier_name = None
    grade = None
    carrier_name = None

    if hasattr(export, "deal") and export.deal_id:
        deal = (
            Deals.objects
            .select_related("supplier", "transport_company")
            .filter(id=export.deal_id)
            .first()
        )
        if deal:
            supplier_name = deal.supplier.name if deal.supplier else None
            grade = deal.grade
            carrier_name = deal.transport_company.name if deal.transport_company else None

    data = {
        "id": export.id,
        "date": export.date.isoformat() if export.date else None,

        "lane": export.lane.name if export.lane else None,
        "lane_id": export.lane_id,

        "hs_code": export.hs_code or "",
        "mode": export.mode,
        "mode_label": export.get_mode_display(),
        "status": export.status,
        "status_label": export.get_status_display(),

        "export_price": str(export.export_price) if export.export_price is not None else None,
        "export_currency": export.export_currency,

        "container_number": export.container_number or "",
        "seal_number": export.seal_number or "",
        "cers_number": export.cers_number or "",

        "etd": export.etd.isoformat() if export.etd else None,
        "eta": export.eta.isoformat() if export.eta else None,

        # derived from deal (optional)
        "supplier": supplier_name,
        "grade": grade,
        "carrier": carrier_name,

        # schedule derived
        "schedule_id": export.schedule_id or None,
        "bkg_number": export.schedule.bkg_number if export.schedule else None,
        "vessel": export.schedule.vessel if export.schedule else None,
        "doc_cutoff_at": export.schedule.doc_cutoff_at.isoformat() if export.schedule and export.schedule.doc_cutoff_at else None,
        "erd_at": export.schedule.erd_at.isoformat() if export.schedule and export.schedule.erd_at else None,
        "cargo_cutoff_at": export.schedule.cargo_cutoff_at.isoformat() if export.schedule and export.schedule.cargo_cutoff_at else None,

        "docs_count": export.documents.count(),
        "documents": [
            {
                "id": d.id,
                "doc_type": d.doc_type,
                "doc_type_label": d.get_doc_type_display(),
                "file_name": d.file.name.split("/")[-1] if d.file else "",
                "file_url": d.file.url if d.file else "",
                "uploaded_at": d.uploaded_at.isoformat() if d.uploaded_at else None,
            }
            for d in export.documents.all()
        ],
    }

    return JsonResponse(data)


# -----------------------------
# EXCEL (stub)
# -----------------------------

def export_shipments_to_excel(request):
    return HttpResponse("Not implemented yet", content_type="text/plain")


# -----------------------------
# CREATE (POST JSON)
# -----------------------------

@require_POST
def export_shipment_create(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    def clean(v): return (v or "").strip()

    # lane (id or name)
    lane_obj = None
    lane_val = payload.get("lane")
    try:
        lane_id = int(lane_val)
        lane_obj = ExportLane.objects.filter(id=lane_id).first()
    except Exception:
        lane_name = clean(lane_val)
        if lane_name:
            lane_obj, _ = ExportLane.objects.get_or_create(name=lane_name)

    # schedule AUTO (from manual inputs)
    bkg = clean(payload.get("bkg_number"))
    vessel = clean(payload.get("vessel"))

    doc_cutoff_at = parse_datetime(payload.get("doc_cutoff_at")) if payload.get("doc_cutoff_at") else None
    erd_at = parse_datetime(payload.get("erd_at")) if payload.get("erd_at") else None
    cargo_cutoff_at = parse_datetime(payload.get("cargo_cutoff_at")) if payload.get("cargo_cutoff_at") else None

    schedule_obj = None
    if bkg or vessel or doc_cutoff_at or erd_at or cargo_cutoff_at:
        # создаём отдельный schedule под конкретную отгрузку (это твой кейс)
        schedule_obj = VesselSchedule.objects.create(
            lane=lane_obj if lane_obj else ExportLane.objects.first(),  # если lane пустой — подстраховка
            bkg_number=bkg or "-",
            vessel=vessel or "-",
            doc_cutoff_at=doc_cutoff_at,
            erd_at=erd_at,
            cargo_cutoff_at=cargo_cutoff_at,
            is_active=True,
        )

    export = ExportShipment.objects.create(
        date=parse_date(payload.get("date")) if payload.get("date") else None,
        lane=lane_obj,
        schedule=schedule_obj,

        hs_code=clean(payload.get("hs_code")),
        mode=clean(payload.get("mode")) or ExportShipment.MODE_OCEAN,
        status=clean(payload.get("status")) or ExportShipment.STATUS_DRAFT,

        export_price=payload.get("export_price") or None,
        export_currency=clean(payload.get("export_currency")) or "USD",

        container_number=clean(payload.get("container_number")),
        seal_number=clean(payload.get("seal_number")),
        cers_number=clean(payload.get("cers_number")),

        etd=parse_date(payload.get("etd")) if payload.get("etd") else None,
        eta=parse_date(payload.get("eta")) if payload.get("eta") else None,

        created_by=request.user if request.user.is_authenticated else None,
    )

    return JsonResponse({"ok": True, "id": export.id})


# -----------------------------
# UPDATE (PATCH or POST JSON)
# -----------------------------

@require_http_methods(["PATCH", "POST"])
def export_shipment_update(request, pk):
    export = get_object_or_404(ExportShipment.objects.select_related("lane", "schedule"), pk=pk)

    try:
        payload = json.loads((request.body or b"{}").decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    # lane
    if "lane" in payload:
        export.lane = _lane_from_payload(payload.get("lane"))

    # schedule FK
    if "schedule" in payload:
        export.schedule = _schedule_from_payload(payload.get("schedule"))

    # deal (only if exists)
    if hasattr(ExportShipment, "deal") and "deal" in payload:
        export.deal = _deal_from_payload(payload.get("deal"))

    # простые строки на самом ExportShipment
    for field in ["hs_code", "mode", "container_number", "seal_number", "cers_number", "export_currency"]:
        if field in payload:
            setattr(export, field, _clean_str(payload.get(field)))

    if "export_price" in payload:
        val = payload.get("export_price")
        export.export_price = None if val in ("", None) else val

    for field in ["date", "etd", "eta"]:
        if field in payload:
            setattr(export, field, _parse_date_or_none(payload.get(field)))

    export.save()

    # ── ОБНОВЛЯЕМ VesselSchedule ──────────────────────────────────────────────
    schedule_fields = {"vessel", "bkg_number", "doc_cutoff_at", "erd_at", "cargo_cutoff_at"}
    schedule_payload = {k: payload[k] for k in schedule_fields if k in payload}

    if schedule_payload:
        if export.schedule_id:
            # schedule уже привязан — обновляем его
            sch = export.schedule
        else:
            # schedule нет — создаём новый и привязываем
            sch = VesselSchedule(
                lane=export.lane or ExportLane.objects.first(),
                bkg_number="-",
                vessel="-",
                is_active=True,
            )
            # сохраним после заполнения ниже

        for field in ["vessel", "bkg_number"]:
            if field in schedule_payload:
                setattr(sch, field, _clean_str(schedule_payload[field]) or "-")

        for field in ["doc_cutoff_at", "erd_at", "cargo_cutoff_at"]:
            if field in schedule_payload:
                setattr(sch, field, _parse_dt_or_none(schedule_payload[field]))

        sch.save()

        if not export.schedule_id:
            export.schedule = sch
            export.save(update_fields=["schedule"])
    # ─────────────────────────────────────────────────────────────────────────

    return JsonResponse({
        "ok": True,
        "id": export.id,
        "lane":            export.lane.name if export.lane else "",
        "lane_id":         export.lane_id or "",
        "schedule_id":     export.schedule_id or "",
        "bkg_number":      export.schedule.bkg_number if export.schedule else "",
        "vessel":          export.schedule.vessel if export.schedule else "",
        "doc_cutoff_at":   export.schedule.doc_cutoff_at.isoformat() if export.schedule and export.schedule.doc_cutoff_at else "",
        "erd_at":          export.schedule.erd_at.isoformat() if export.schedule and export.schedule.erd_at else "",
        "cargo_cutoff_at": export.schedule.cargo_cutoff_at.isoformat() if export.schedule and export.schedule.cargo_cutoff_at else "",
        "status":          export.status,
        "mode":            export.mode,
        "hs_code":         export.hs_code,
        "export_price":    str(export.export_price) if export.export_price is not None else "",
        "export_currency": export.export_currency,
    })


# -----------------------------
# INLINE SINGLE-FIELD UPDATE (POST JSON)
# -----------------------------

@require_POST
def export_shipment_update_field(request, pk):
    export = get_object_or_404(ExportShipment, pk=pk)

    try:
        data = json.loads((request.body or b"{}").decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON"}, status=400)

    field = data.get("field")
    value = data.get("value")

    allowed = {
        "date",
        "lane",
        "schedule",
        "hs_code",
        "mode",
        "export_price",
        "export_currency",
        "container_number",
        "seal_number",
        "cers_number"
        "etd",
        "eta",
    }

    if field not in allowed:
        return JsonResponse({"ok": False, "error": f"Field not allowed: {field}"}, status=400)

    if field == "lane":
        export.lane = _lane_from_payload(value)
    elif field == "schedule":
        export.schedule = _schedule_from_payload(value)
    elif field in ("date", "etd", "eta"):
        setattr(export, field, _parse_date_or_none(value))
    elif field == "export_price":
        export.export_price = None if value in ("", None) else value
    else:
        setattr(export, field, _clean_str(value))

    export.save()
    return JsonResponse({"ok": True, "id": export.id})


@require_POST
def export_shipment_upload(request, pk):
    export = get_object_or_404(ExportShipment, pk=pk)

    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return JsonResponse({"ok": False, "error": "No file provided"}, status=400)

    doc_type = (request.POST.get("doc_type") or "").strip() or "other"

    doc = ExportShipmentDocument.objects.create(
        export=export,
        file=uploaded_file,
        doc_type=doc_type,
        uploaded_by=request.user if hasattr(ExportShipmentDocument, "uploaded_by") else None,
    )

    return JsonResponse({
        "ok": True,
        "doc": {
            "id": doc.id,
            "doc_type": getattr(doc, "doc_type", ""),
            "doc_type_label": doc.get_doc_type_display() if hasattr(doc, "get_doc_type_display") else getattr(doc, "doc_type", ""),
            "file_name": doc.file.name.split("/")[-1] if doc.file else "",
            "file_url": doc.file.url if doc.file else "",
            "uploaded_at": doc.uploaded_at.isoformat() if getattr(doc, "uploaded_at", None) else None,
        }
    })

@require_POST
def export_document_upload(request, pk):
    export = get_object_or_404(ExportShipment, pk=pk)

    f = request.FILES.get("file")
    if not f:
        return JsonResponse({"ok": False, "error": "No file provided"}, status=400)

    doc_type = (request.POST.get("doc_type") or "other").strip()
    allowed = {c[0] for c in ExportDocument.DOC_TYPE_CHOICES}
    if doc_type not in allowed:
        doc_type = "other"

    doc = ExportDocument.objects.create(
        export_shipment=export,
        doc_type=doc_type,
        file=f,
    )

    return JsonResponse({
        "ok": True,
        "doc": {
            "id": doc.id,
            "doc_type": doc.doc_type,
            "doc_type_label": doc.get_doc_type_display(),
            "file_name": doc.file.name.split("/")[-1] if doc.file else "",
            "file_url": doc.file.url if doc.file else "",
            "uploaded_at": doc.uploaded_at.isoformat() if doc.uploaded_at else None,
        }
    })

@require_POST
def export_shipment_delete(request, pk):
    export = get_object_or_404(ExportShipment, pk=pk)
    export.delete()
    return JsonResponse({"ok": True, "id": pk})

# -----------------------------
# Scale tickets mass archive
# -----------------------------


@csrf_exempt
@require_POST
def generate_current_month_scale_tickets_archive(request):
    try:
        payload = json.loads(request.body or "{}")
    except Exception:
        payload = {}

    month = payload.get("month")
    year = payload.get("year")

    today = t.localdate()

    try:
        month = int(month) if month else today.month
        year = int(year) if year else today.year
    except ValueError:
        return JsonResponse({
            "success": False,
            "error": "Invalid month or year"
        }, status=400)

    deals = (
        Deals.objects
        .filter(date__year=year, date__month=month)
        .exclude(scale_ticket__isnull=True)
        .exclude(scale_ticket__exact="")
        .select_related("supplier")
        .order_by("scale_ticket", "date", "id")
    )

    unique_ticket_numbers = []
    seen = set()

    for deal in deals:
        ticket_number = str(deal.scale_ticket).strip()
        if not ticket_number or ticket_number in seen:
            continue
        seen.add(ticket_number)
        unique_ticket_numbers.append(ticket_number)

    created = 0
    updated = 0
    errors = []

    rf = RequestFactory()

    for ticket_number in unique_ticket_numbers:
        try:
            ticket_deals = (
                Deals.objects
                .filter(scale_ticket=ticket_number, date__year=year, date__month=month)
                .select_related("supplier")
                .order_by("date", "id")
            )

            if not ticket_deals.exists():
                continue

            first_deal = ticket_deals.first()

            if not first_deal.supplier:
                errors.append(f"Ticket {ticket_number}: supplier not found")
                continue

            date_src = first_deal.date or today
            file_year = date_src.strftime("%Y")
            month_num = date_src.strftime("%m")
            month_name = date_src.strftime("%B")
            month_dir = f"{file_year}-{month_num}"

            raw_supplier_name = first_deal.supplier.name if first_deal.supplier else "Unknown Supplier"
            supplier_name = sanitize_filename(raw_supplier_name)

            filename = f"Ticket {ticket_number}-{supplier_name}-{month_dir}.pdf"

            directory = os.path.join(
                settings.MEDIA_ROOT,
                "reports",
                "scale_tickets",
                supplier_name,
                file_year,
                month_name
            )

            filepath = os.path.join(directory, filename)

            file_existed = os.path.exists(filepath)

            if file_existed:
                os.remove(filepath)

            os.makedirs(directory, exist_ok=True)

            internal_request = rf.get(
                "/fake-export-scale-ticket/",
                {
                    "ticket_number": str(ticket_number),
                    "licence_plate": "N/A",
                    "tare_weight": "5170",
                    "time": datetime.now().strftime("%H:%M"),
                }
            )

            response = export_scale_ticket_pdf(internal_request)

            if response.status_code != 200:
                errors.append(f"Ticket {ticket_number}: PDF generation failed with status {response.status_code}")
                continue

            if os.path.exists(filepath):
                if file_existed:
                    updated += 1
                else:
                    created += 1
            else:
                errors.append(f"Ticket {ticket_number}: file was not created")

        except Exception as e:
            errors.append(f"Ticket {ticket_number}: {str(e)}")

    return JsonResponse({
        "success": True,
        "created": created,
        "updated": updated,
        "errors": errors,
        "month": month,
        "year": year,
    })

def supplier_shipment_report_archive(request):
    return render(request, "crm/supplier_shipment_report_archive.html")


@csrf_exempt
@require_POST
def generate_current_month_supplier_shipment_report_archive(request):
    try:
        payload = json.loads(request.body or "{}")
    except Exception:
        payload = {}

    month = payload.get("month")
    year = payload.get("year")

    today = t.localdate()

    try:
        month = int(month) if month else today.month
        year = int(year) if year else today.year
    except ValueError:
        return JsonResponse({
            "success": False,
            "error": "Invalid month or year"
        }, status=400)

    try:
        generated_files = generate_supplier_shipment_reports_for_month(year, month)

        return JsonResponse({
            "success": True,
            "message": f"Generated {len(generated_files)} supplier shipment report files",
            "count": len(generated_files),
            "year": year,
            "month": str(month).zfill(2),
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        }, status=500)


def supplier_shipment_report_archive_years(request):
    base_dir = os.path.join(settings.MEDIA_ROOT, "supplier_shipment_reports")
    years = []

    if os.path.exists(base_dir):
        for name in os.listdir(base_dir):
            path = os.path.join(base_dir, name)
            if os.path.isdir(path) and name.isdigit():
                years.append(name)

    years.sort(reverse=True)

    return JsonResponse({
        "success": True,
        "years": years
    })


def supplier_shipment_report_archive_months(request, year):
    year_dir = os.path.join(settings.MEDIA_ROOT, "supplier_shipment_reports", str(year))
    months = []

    if os.path.exists(year_dir):
        for name in os.listdir(year_dir):
            path = os.path.join(year_dir, name)
            if os.path.isdir(path):
                months.append(name)

    months.sort(reverse=True)

    return JsonResponse({
        "success": True,
        "months": months
    })


def supplier_shipment_report_archive_files(request, year, month):
    month_str = str(month).zfill(2)
    month_dir = os.path.join(
        settings.MEDIA_ROOT,
        "supplier_shipment_reports",
        str(year),
        month_str
    )

    files = []

    if os.path.exists(month_dir):
        for name in os.listdir(month_dir):
            if name.lower().endswith(".pdf"):
                files.append({
                    "name": name,
                    "url": f"{settings.MEDIA_URL}supplier_shipment_reports/{year}/{month_str}/{name}"
                })

    files.sort(key=lambda x: x["name"].lower())

    return JsonResponse({
        "success": True,
        "files": files
    })